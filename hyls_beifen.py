#-*-coding:utf-8-*-


import pandas as pd
import numpy
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill, Alignment
from openpyxl.styles import colors
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Series, Reference,PieChart,LineChart,PieChart3D,ProjectedPieChart
from openpyxl.chart.shapes import GraphicalProperties,LineProperties
from csv_to_excel import csv_to_xlsx, skip1_r
# from openpyxl.chart.series import DataPoint
import time
import openpyxl
# from win32com.client import DispatchEx, constants, gencache,Dispatch


from texttable import Texttable
import os
import xlrd
import xlwt

from pylab import *                                 #字体设置
mpl.rcParams['font.sans-serif'] = ['SimHei']
from pandas import DataFrame
import matplotlib.pyplot as plt
import xlsxwriter
from copy import deepcopy

import datetime
from datetime import timedelta
# from datetime import datetime
start_time = time.time()

# import datetime as dt
from dateutil.parser import parse

#
# dizhi = r'/Users/win11/Desktop/2.3银行流水'
dizhi1 = r'/Users/hh/Desktop/2.3银行流水/银行流水分析'


result = pd.ExcelWriter(r'/Users/hh/Desktop/2.3银行流水/1.xlsx')
result1 = pd.ExcelWriter(r'/Users/hh/Desktop/2.3银行流水/2.xlsx')
result2 = pd.ExcelWriter(r'/Users/hh/Desktop/2.3银行流水/3.xlsx')
result3 = pd.ExcelWriter(r'/Users/hh/Desktop/2.3银行流水/6.xlsx')
result4 = pd.ExcelWriter(r'8.xlsx', engine='xlsxwriter')
result9 = pd.ExcelWriter(r'/Users/hh/Desktop/2.3银行流水/888.xlsx')
result5 = pd.ExcelWriter(r'9.xlsx',engine='xlsxwriter')
# glqy_list = '关联企业-人为峰'
company_name = 0
#获取表格列表

"""
#1 文件拼接
"""

def skip_r(file_path):

    content = pd.read_excel(file_path, skiprows=0, skipfooter=0)
    col = content.columns
    col = list(col)
    a = 0
    if '摘要' in col or '附言' in col or '交易信息' in col or '交易用途' in col:
        pass
    else:
        for i in range(0, 50):
            a = 0
            cont_row = content.iloc[i, :].to_list()
            for j in cont_row:
                j = str(j)
                if j.find('摘要') >= 0 or j.find('附言') >= 0 or j.find('交易信息') >= 0 or j.find('交易用途') >= 0:
                    a = i + 1
                    # return a
                    break
            if a != 0:
                break
        return a



def skip_f(file_path,m):
    content = pd.read_excel(file_path, skiprows=m, skipfooter=0)
    idx = content.tail(1)
    foot = idx.iloc[0, len(idx.columns) - 1]
    b = 0
    if foot is not nan:
        pass
    else:
        for i in range(0, 50):
            b += 1
            content = pd.read_excel(file_path, skiprows=m, skipfooter=b)
            idx = content.tail(1)
            foot = idx.iloc[0, len(idx.columns) - 1]
            if foot is not nan:
            # if len(idx.columns)>8:
                break
    return b


def rd_data(a, ls):
    ls0 = pd.DataFrame()
    excel_amount = len(origin_file_list)
    excel_df = []*excel_amount
    # print('aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa123')
    for i, value in enumerate(origin_file_list):
        # 拼接每个文件的路径
        file_path = '/Users/hh/Desktop/2.3银行流水/银行流水分析/%s/%s' % (a, value)
        f = pd.ExcelFile(file_path)
        f_sheet = f.sheet_names
        # print(len(f_sheet))
        if len(f_sheet)>1:
            # if value =='.DS_Store' :
        #     pass
            m = skip_r(file_path)
            n = skip_f(file_path, m)
            # print(m,n)
            # print('woaini')
            for i in f.sheet_names:
                # print(i)
                data = pd.read_excel(file_path, skiprows=m, skipfooter=n, sheet_name=i)
                content = pd.concat([data, content])

        elif file_path.find('xls') >= 0:
            m = skip_r(file_path)
            n = skip_f(file_path, m)
            content = pd.read_excel(file_path, skiprows=m, skipfooter=n)
            # data = pd.concat([data, content])

        elif file_path.find('csv') >= 0:
            content = csv_to_xlsx(file_path)
        # content = pd.read_excel(file_path, skiprows=m, skipfooter=n)
        if ' ' in content.columns:
                # row0_list = ls.iloc[0,:].to_list()
                # print(ls)
                # print(ls.loc[0, ' '])
            row0_1 = content.loc[0, ' ']
            if row0_1.find('贷') >= 0 or row0_1.find('收') >= 0:
                content.rename(columns={' ': '收入'}, inplace=True)
                content.rename(columns={'发生额/元': '支出'}, inplace=True)
            elif row0_1.find('借') >= 0 or row0_1.find('支') >= 0:
                content.rename(columns={' ': '支出'}, inplace=True)
                content.rename(columns={'发生额/元': '收入'}, inplace=True)

            content = content.drop(index=[0])
        # excel_df[i] =content
        ls0 = pd.concat([ls0,content],axis=0,join='outer')
    ls0 = ls0.reset_index(drop=True)
    # print(m)
    return ls0, m


def strange_float(h=[]):
    g = []
    for i in h:
        i = str(i)
        if ',' in i:
            a = i.replace(',', '')
            b = float(a)
            g.append(b)
            h = g
        elif ('-' in i) or (' ' in i) or i == '':
            i = 0.0
            g.append(i)
            h = g
        else:
            i = float(i)
            g.append(i)
            h = g
    return h


"""
交易日期与交易时间合并
"""

def strange_time(ls):   ####合并交易日期与交易时间

    col_list = []
    for i in ls.columns:
        i = str(i)
        a = i.replace(" ", "")
        col_list.append(a)
    ls.columns = col_list
    for i in ls.columns:
        if i.find('交易时间')>=0:
            ls.rename(columns={i: '交易时间'}, inplace=True)
        elif i.find('日期')>=0:
            ls.rename(columns={i: '交易日期'}, inplace=True)

    if '交易日期' in ls.columns and '交易时间' in ls.columns:
        a = ls.loc[8,'交易日期']
        b = ls.loc[8,'交易时间']
        if str(a) not in str(b):
            if type(ls.loc[8,'交易日期']) is pd._libs.tslibs.timestamps.Timestamp:
                ls['交易日期'] = ls['交易日期'].apply(lambda x: datetime.datetime.strftime(x, '%Y-%m-%d'))

            ls['交易日期'] = ls['交易日期'].apply(lambda x: str(x))
            ls['交易时间'] = ls['交易时间'].apply(lambda x: str(x))
            ls['交易时间'] = ls['交易日期'] + ls['交易时间']
            try:

            # if ls.loc[5,'交易时间'].find(':')>0:
                ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d%H:%M:%S'))
            except Exception:
                pass
            try:
            # elif ls.loc[5,'交易时间'].find(':')<0:
                ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d%H%M%S'))
            except Exception:
                pass
            try:
                # elif ls.loc[5,'交易时间'].find(':')<0:
                ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d%H:%M:%S'))
                print('akakkakakakakakaakkaakakkakakakakakakkakakaka47')
            except Exception:
                pass
            try:
                # elif ls.loc[5,'交易时间'].find(':')<0:
                ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d%H%M%S'))
            except Exception:
                pass
            try:
                # elif ls.loc[5,'交易时间'].find(':')<0:
                ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d%H:%M'))
            except Exception:
                pass
        else:
            pass
    elif '交易日期' in ls.columns and '交易时间' not in ls.columns:
        ls.rename(columns={'交易日期': '交易时间'}, inplace=True)
    return ls

# for i in range(0,ls_amount):
#     ls_yhls[i]=strange_time(ls_yhls[i])


"""
#3 时间格式转换
"""

def strange_time1(ls):

    a = ls.loc[4,'交易时间']
    if type(ls.loc[4, '交易时间']) is pd._libs.tslibs.timestamps.Timestamp :
        pass
    elif type(ls.loc[4, '交易时间']) is datetime.time:
        pass
    elif type(ls.loc[4, '交易时间']) is numpy.int64:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: str(x))
    # elif a.find('\\') < 0 and a.find('-') < 0 and len(a) >= 20:  # 语法没错
    #
    #     ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x[0:14], '%Y%m%d%H%M%S'))
    #
    # elif a.find('\\') < 0 and a.find('-') < 0 and a.find(':') < 0 and len(a) == 8:  # 语法没错
    #
    #     ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d'))
    #
    # elif a.find('-') > 0 and a.find(':') > 0:
    #     ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
    #
    #     print('abc')
    # elif a.find(':') > 0 and len(a) > 8:
    #     ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d %H:%M:%S'))
    #     print('abcd')
    # elif a.find('\\') > 0:
    #
    #     ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y\%m\%d %H\%M\%S'))
    # else:
    #     pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x[0:14], '%Y%m%d%H%M%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y/%m/%d'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y%m%d %H:%M:%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y\%m\%d %H\%M\%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y/%m/%d %H\%M\%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y/%m/%d %H:%M'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y/%m/%d %H:%M:%S'))
    except Exception:
        pass
    try:
        ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y\%m\%d %H:%M'))
    except Exception:
        pass
    return ls

# for i in range(0,ls_amount):
#     ls_yhls[i]=strange_time1(ls_yhls[i])

"""
#2 统一格式
"""
# a=0
zhanghu_list=[]
def strange_zhanghu(ls):
    global ind_name
    a=0
    b=0
    c=0
    for i in ls.columns:
        if i.find('对方户名')>=0:
            a=1
            break
        elif i.find('收款人名称')==0:
            a=2
            b=i
            # break
        elif i.find('付款人名称',0)==0:
            c=i

    if a == 2:
        ls1 = ls.groupby(b).count()
        col_name = ls1.columns[0]
        ls1 = ls1.sort_values(by=col_name, ascending=True)
        ls1= ls1.tail(1)
        ind_name = ls1.index[0]

        for i,value in enumerate(ls[b]):
            value =str(value)
            if value ==ind_name:
                ls.loc[i,b] = ls.loc[i,c]
        ls.rename(columns={b: '对方户名'}, inplace=True)
    # ls.loc[ind_name, b] = ls.loc[ind_name, c]
    return ls

# for i in range(0,ls_amount):
#     ls_yhls[i] = strange_zhanghu(ls_yhls[i])


def strange_geshi(ls):

    if ls is None:
        pass
    else:
        ls_gs = pd.DataFrame(columns=['交易时间', '对方户名', '收入', '支出', '余额', '摘要', '附言', '备注', '交易信息', '交易用途'])
        ls_item = ls.columns
        ls_item = list(ls_item)
        # print(ls_item)
        for i in ls_item:
            if i.find('交易金额') >= 0:
                ls.rename(columns={i: '交易金额'}, inplace=True)
        #         ls_je = ls_no['金额'].tolist()
        #         if type(ls_je[7]) is not float:
        #             ls_je = strange_float(ls_je)
        #             ls_no['金额'] = pd.Series(ls_je)
        #         break
        js = 0
        for i in ls_item:

            if i.find('贷方') >= 0 or i.find('收入') >= 0:
                break
            else:
                js += 1

        if js == len(ls_item):
            ls['支出'] = ls['交易金额']
            ls.loc[ls[(ls['支出'] > 0)].index, ['支出']] = 0
            ls['支出'] = ls['支出'].apply(lambda x: x*(-1))
            ls.rename(columns={'交易金额': '收入'}, inplace=True)
            ls.loc[ls[(ls['收入'] < 0)].index, ['收入']] = 0
            ls_item = ls.columns
            ls_item = list(ls_item)
            for i in ls_item:

                if i == '对方账号户名' or i == '对方户名' or i == '对方单位名称' or i == '对方名称' or i == '对方账号名称':
                    ls.rename(columns={i: '对方户名'}, inplace=True)
                elif i.find('交易时间') > 0 or i == '日期':
                    ls.rename(columns={i: '交易时间'}, inplace=True)
                elif i.find('上笔') >= 0:
                    ls.rename(columns={i: '上笔'}, inplace=True)
                elif i.find('余额') > 0 or i.find('账户余额') > 0:
                    ls.rename(columns={i: '余额'}, inplace=True)
                elif i.find('附言') >= 0:
                    ls.rename(columns={i: '附言'}, inplace=True)
                elif i.find('摘要') >= 0:
                    ls.rename(columns={i: '摘要'}, inplace=True)
                elif i.find('用途') >= 0:
                    ls.rename(columns={i: '交易用途'}, inplace=True)
                else:
                    pass
        else:

            for i in ls_item:
                if i=='对方账号户名' or i == '对方户名' or i == '对方名称' or i == '对方单位名称' or i == '对方账号名称':
                    ls.rename(columns={i: '对方户名'}, inplace=True)
                elif i.find('交易时间') >= 0 or i == '日期':
                    ls.rename(columns={i: '交易时间'}, inplace=True)
                elif i.find('上笔') >= 0:
                    ls.rename(columns={i: '上笔'}, inplace=True)
                elif i.find('余额') >= 0 or i.find('账户余额') >= 0:
                    ls.rename(columns={i: '余额'}, inplace=True)
                elif i.find('贷方') >= 0 or i.find('收') >= 0:
                    ls.rename(columns={i: '收入'}, inplace=True)
                elif i.find('借方') >= 0 or i.find('支') >= 0:
                    ls.rename(columns={i: '支出'}, inplace=True)
                elif i.find('附言') >= 0:
                    ls.rename(columns={i: '附言'}, inplace=True)
                elif i.find('摘要') >= 0:
                    ls.rename(columns={i: '摘要'}, inplace=True)
                elif i.find('用途') >= 0:
                    ls.rename(columns={i: '交易用途'}, inplace=True)
                else:
                    pass

        ls = pd.concat([ls_gs, ls], axis=0, join='inner')
        ls = ls.reset_index(drop=True)
        return ls

# for i in range(0, ls_amount):
#     ls_yhls[i] = strange_geshi(ls_yhls[i])


"""
#4 数字类型转换（去掉'，'）
"""
def strange_datatype(ls):

    ls['余额'].fillna(0, inplace=True)
    ls['收入'].fillna(0, inplace=True)
    ls['支出'].fillna(0, inplace=True)
    # ls['余额'].fillna('0', inplace=True)
    # ls['收入'].fillna('0', inplace=True)
    # ls['支出'].fillna('0', inplace=True)
    ls_da1 = ls['余额'].tolist()
    ls_da2 = ls['收入'].tolist()
    ls_da3 = ls['支出'].tolist()
    # if type(ls_da1[5]) is not float:
    ls_da1 = strange_float(ls_da1)
    ls['余额'] =pd.Series(ls_da1)
    # if type(ls_da2[5]) is not float:
    ls_da2 = strange_float(ls_da2)
    ls['收入'] =pd.Series(ls_da2)
    # if type(ls_da3[5]) is not float:
    ls_da3 = strange_float(ls_da3)
    ls['支出'] =pd.Series(ls_da3)
    # else:
    return ls


# for i in range(0,ls_amount):
#     ls_yhls[i] = strange_datatype(ls_yhls[i])



"""
判断 是否有附言
"""
def strange_zhaiyao(ls):

    js1=0
    for i in ls.columns:
        if i.find('附言')>=0:
            a=i
            break
        elif i.find('摘要')>=0:
            a=i
            # break
        elif i.find('交易信息')>=0:
            a=i
        elif i.find('备注')>=0:
            a=i
        elif i.find('交易用途') >= 0:
            a=i
    for i in ls.columns:
        if i.find('摘要') >= 0 or i.find('用途') >= 0 or i.find('交易信息') >= 0 or i.find('备注') >= 0 or i.find('附言')>=0:
            if a !=i:
                ls[i] = ls[i].apply(lambda x: str(x))
                ls[a] = ls[a].apply(lambda x: str(x))
                ls[a] = ls[a]+ls[i]
                ls = ls.drop(i, axis=1)
    if a != '附言':
        ls.rename(columns={a: '附言'}, inplace=True)
        # ls = ls.drop(i, axis=1)
    # for i in ls.columns:
    #     if i.find('附言') >= 0:
    #         pass
    #     elif i.find('摘要') >= 0 or i.find('交易信息') >= 0 or i.find('备注') >= 0:
    #         ls.rename(columns={i: '附言'}, inplace=True)
    return ls



def strange_index(ls):
    ls = strange_zhaiyao(ls)
    print(ls)
    ls.set_index('交易时间', inplace=True)
    print(ls)
    return ls


# ls_yhls1 = ls_amount*[' ']
# # for i in range(0,ls_amount):
# #     ls_yhls1[i] = setind(ls_yhls[i])
#
#
# ls = pd.DataFrame()
# for i in range(0,ls_amount):
#     # if i <= ls_amount:
#     ls_yhls[i] = setind(ls_yhls[i])
#     ls=pd.concat([ls,ls_yhls[i]], axis =0 ,join='outer')
#     ls_yhls1[i] = ls_yhls[i]


# ls_yhls[0].to_excel(result4 , sheet_name='银行流水1' , index=True)
# ls_yhls[1].to_excel(result4 , sheet_name='银行流水2' , index=True)
# ls_yhls[2].to_excel(result4 , sheet_name='银行流水3' , index=True)
# ls_yhls[3].to_excel(result4 , sheet_name='银行流水4' , index=True)
# result4.save()


# for i in ls['交易时间']:
# ls['交易时间'] = ls['交易时间'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
# ls.set_index('交易时间', inplace=True)


###### ls = ls.sort_values(by=['交易时间'], ascending=True)
##### ls.to_excel(result1, sheet_name='银行流水0', index=True)
###### result1.save()


# ls.set_index('交易时间', inplace=True)
# ls_fenweishu= ls.resample('D',kind='period').sum()
# ls_fenweishu['金额'] = ls_fenweishu['收入'] + ls_fenweishu['支出']
# ls_fws = ls_fenweishu.describe()
# fws = ls_fws.loc['75%', '金额']
def file_name(file_dir):
    excel_list1 = []
    for root, dirs, files in os.walk(file_dir):
        # print(root) #当前目录路径
        # print(dirs) #当前路径下所有子目录
        print(files) #当前路径下所有非目录子文件

        excel_list1.append(files)
        print(excel_list1)
        print('aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa')
    return excel_list1

excel_list1 = file_name(dizhi1)

if ['.DS_Store'] in excel_list1:
    del excel_list1[0]








origin_file = os.listdir(r'%s' % (dizhi1))
print(origin_file)
if '.DS_Store' in origin_file:
    del origin_file[0]    ##删除.DS_store
# print(origin_file)
print('iiiii')

excel_list = []
for i in range(0, len(origin_file)):
    origin_file_list = os.listdir(r'%s/%s' % (dizhi1, origin_file[i]))
    if '.DS_Store' in origin_file_list:
        del origin_file_list[0]
    if len(origin_file_list):  ##去掉空文件夹
        excel_list.append(origin_file_list)
    # print()
    print(excel_list)

ls_amount = len(excel_list)

a=0
ls_yhls = ['']*ls_amount
ls_yhls1 = ['']*ls_amount
for i in excel_list:
    print(i)
    ls0 = pd.DataFrame()
    for j in i:
        print(j)
        file_path = '/Users/hh/Desktop/2.3银行流水/银行流水分析/%s/%s' % (a, j)
        m = skip_r(file_path)
        n = skip_f(file_path, m)
        ls = pd.read_excel(file_path, skiprows=m, skipfooter=n)
        ls = strange_time(ls)
        ls = strange_time1(ls)
        ls = strange_zhanghu(ls)
        ls = strange_geshi(ls)
        ls = strange_datatype(ls)
        ls = strange_index(ls)
        ls0 = pd.concat([ls0,ls],axis=0,join='outer')
        print(ls)
    ls_yhls[a] = ls0
    ls_yhls1[a] = ls0
    # ls1 = pd.read_excel(r'/Users/win11/Desktop/2.3银行流水/2.xlsx', skiprows=0)
    a+=1

ls0.to_excel(result1, sheet_name='银行流水0', index=True)
result1.save()

print('i love you')
print(ls_yhls[0])
print('pppppppppppppppppppppppp123')
print('aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa')


skip_rows = []
# for i in range(0, 15):
#
#     if a < ls_amount:
#         ls_yhls[a] = rd_data('%s'%(a), excel_list[a])[0] ###超市工行-3237账户    光大18年   工行
#         skip_row = rd_data('%s'%(a), excel_list[a])[1]
#         # print(skip_row)
#         skip_rows.append(skip_row)
#         a += 1
#     else:
#         break

try:
    for i in skip_rows:
        if i>1:
            ind0 = skip_rows.index(i)
            front = i
            break

    ls_company = pd.read_excel(r'%s/%s/%s'%(dizhi1, ind0, excel_list[ind0][0]))
    # print(ls_company)
    for i in range(0, front-1):

        ls_company_front = ls_company.iloc[i, :].to_list()
        for j in ls_company_front:
            j=str(j)
            if j.find('公司') >= 0:
                company_name = j
                break

except Exception:
    pass

"""
#5 数据分析
"""
len0 = 15      #标题15行
len_gp = 3      #模块间距3行
len_pic = 20    #图片20行
len_rjye = 60    #日均余额绝对位置
len_rjye1 = len_rjye+20


"""
时间columns 格式转换
"""
def strange_col(ls_item):

    date_cl = []
    ls_tt = ls_item.columns
    for i in ls_tt:
        a = str(i)
        # print(type(a))
        date_cl.append(a)
    # print(date_cl)
    ls_item.columns = date_cl

"""
日均余额模块
"""
def rjye(ls):


    ls_rjye = ls['余额'].resample('D').last()

    ls_rjye = ls_rjye.fillna(method='ffill')



    print(ls_rjye)
    print('pppppppppppppppppppppppppppppppppp12333')

    ls_rjye = ls_rjye.resample('M',kind='period').mean()
    #
    # ls_rjye = ls_rjye.T
    return ls_rjye

ls_yhls2 = ls_amount*[' ']
for i in range(0,ls_amount):
    ls_yhls2[i] = rjye(ls_yhls1[i])
    ls_yhls2[i] =pd.Series(ls_yhls2[i])
# ls_nh2 = rjye(ls_nh1)
# ls_nht2 = rjye(ls_nht1)
# ls_gh2 = rjye(ls_gh1)
# ls_jh2 = rjye(ls_jh1)



# ls_nh2=pd.Series(ls_nh2)
# ls_nht2=pd.Series(ls_nht2)
# ls_gh2=pd.Series(ls_gh2)
# ls_jh2=pd.Series(ls_jh2)
df = pd.Series()
if ls_amount>=1:
    for i in range(0,ls_amount):
        df = df.add(ls_yhls2[i], fill_value=0.00)



def rjzx(ls):
    if ls is None:
        pass
    else:

        ls_rjzx = ls['余额'].resample('D').last()
        ls_rjzx = ls_rjzx.fillna(method='ffill')
        return ls_rjzx

ls_yhls3 = ls_amount*[' ']
for i in range(0,ls_amount):

    ls_yhls3[i] =pd.Series(rjzx(ls_yhls1[i]))



df1 = pd.Series()
if ls_amount>=1:
    for i in range(0,ls_amount):
        df1 = df1.add(ls_yhls3[i],fill_value=0.00)


df1 = df1.resample('M',kind='period').min()


# ls3[len(ls3)-1]
rjye_qujian = len(df)
print(len(df))
print('ooooooooooooooooooooooooooooooooooooo')
a1 = df[len(df)-1]
a2 = df.tail(3).mean()
a3 = df.tail(6).mean()
if rjye_qujian >= 12:
    a4 = df.tail(12).mean()

b1 = df1[len(df1)-1]
b2 = df1.tail(3).min()
b3 = df1.tail(6).min()
if rjye_qujian >= 12:
    b4 = df1.tail(12).min()

rjzx_my = pd.DataFrame(df1)
rjzx_my = rjzx_my.T
rjzx_my.index=['最小余额']        ###每月日均最小


rjye_my = pd.DataFrame(df)

rjye_my = rjye_my.T
rjye_my.index=['日均余额']        ###每月日均余额


rjye = pd.concat([rjye_my,rjzx_my],axis=0,join='inner')


rjye1 = rjye.T      ###画图使用

# rjye1 = rjye1[rjye1.index.apply(lambda x: datetime.datetime.strftime(x,'%Y-%m'))]

rjye_list=rjye1.index.to_list()

rjye_list = list(map(str, rjye_list))

rjye1.index = rjye_list
len_ye = len(rjye1['日均余额'])


a = len(rjye.columns)

if a < 10:
    rjye.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 7, startcol=1, index=True)
    loc_qs = len_rjye1 +7 +7
    # loc_part4 = loc_qs + 31
    loc_part4 = 153
    loc_part5 = loc_part4 + 69
    len_rj = 3
    len_rj1 = len(rjye.columns)
    print(len_rj1)
    rj_1=0

elif 10 <= a <= 20:
    rjye_my1 = rjye.iloc[:, :int((a+1)//2)]
    rjye_my2 = rjye.iloc[:, int((a+1)//2):]
    rjye_my1.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 7, startcol=1, index=True)
    rjye_my2.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 10, startcol=1, index=True)
    loc_qs = len_rjye1  + 10 + 7
    # loc_part4 = loc_qs + 31
    loc_part4 = 153
    loc_part5 = loc_part4 + 81
    len_rj = 6
    len_rj1 = len(rjye.columns)//2
    print(len_rj1)

    rj_1 = len(rjye.columns) % 2


elif 20 < a <= 30:
    rjye_my1 = rjye.iloc[:, : int(a//3)]
    rjye_my2 = rjye.iloc[:, int(a//3):int(2*a//3)]
    rjye_my3 = rjye.iloc[:, int(2*a//3):]
    rjye_my1.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 7, startcol=1, index=True)
    rjye_my2.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 10, startcol=1, index=True)
    rjye_my3.to_excel(result5, sheet_name='银行流水', startrow=len_rjye1 + 13, startcol=1, index=True)
    loc_qs = len_rjye1 + 13 + 7
    loc_part4 = loc_qs + 31
    loc_part5 = loc_part4 + 93
    len_rj = 9
    len_rj1 = len(rjye.columns) % 3

loc_part6 = loc_part5+22
loc_part7 = loc_part6+9


if rjye_qujian >= 12:
    rjye_qj = pd.DataFrame({'交易时间': ['近一个月', '近三个月', '近半年','近一年'], '日均余额': [a1, a2, a3,a4], '最小余额': [b1, b2, b3,b4]})
else:
    rjye_qj = pd.DataFrame({'交易时间': ['近一个月', '近三个月', '近半年'], '日均余额': [a1, a2, a3], '最小余额': [b1, b2, b3]})

rjye_qj = rjye_qj.set_index("交易时间")

rjye_qj1 = rjye_qj.T              ###期间日均余额



rjye_qj1.to_excel(result5, sheet_name='银行流水' , startrow=len_rjye, startcol=1,index=True)

rjye1.to_excel(result5, sheet_name='银行流水' , startrow=len_rjye+3, startcol=8,index=True)

# b=list(range(0,len(ls_out_bn['Col_sum'])))


"""
数据分析
"""

ls1 = pd.read_excel(r'/Users/hh/Desktop/2.3银行流水/2.xlsx', skiprows=0)

def ls_ic_pay(a,ls1):

    ls_byname = ls1[ls1['{0}'.format(a)].values > 0]

    zls = ls1['{0}'.format(a)].sum()

    zbs = ls_byname['余额'].count()

    ls_by_sum=pd.pivot_table(ls_byname, values='{0}'.format(a), columns='对方户名', index='交易时间', aggfunc='sum')

    ls_byname1 = ls_by_sum.resample('M',kind='period').sum()

    ls_byname2 = ls_by_sum.resample('M', kind='period').count()

    ls_byname1 = ls_byname1.T

    ls_byname2 = ls_byname2.T

    ls_byname1['金额'] = ls_byname1.apply(lambda x : x.sum(), axis=1)    ###计算所有对手方的交易金额

    ls_byname2['笔数'] = ls_byname2.apply(lambda x: x.sum(), axis=1)    ###计算所有对方方的交易笔数

    ls_byname1['笔数'] = ls_byname2['笔数']

    ls_byname1 = ls_byname1[ls_byname1['金额'].values != 0]

    ls_byname1 = ls_byname1.sort_values(by=['金额'], ascending=False)

    ls_sr_ten = ls_byname1.iloc[:10,:]

    ls_sr_ten.loc['总计'] = ls_sr_ten.apply(lambda x: x.sum(), axis=0)

    ls_sr_ten.fillna(0, inplace=True)

    ls_byname1.loc['总计'] = ls_byname1.apply(lambda x: round(x.sum(),2), axis=0)

    ls_byname1.fillna(0, inplace=True)

    strange_col(ls_byname1)

    strange_col(ls_sr_ten)

    return ls_byname1,ls_sr_ten,zls,zbs

ls_in_bn = ls_ic_pay('收入',ls1)[0]
ls_out_bn = ls_ic_pay('支出',ls1)[0]

zlr_je = ls_ic_pay('收入',ls1)[2]   ###总流入金额
zlc_je = ls_ic_pay('支出',ls1)[2]   ###总流出金额


zlr_bs = ls_ic_pay('收入',ls1)[3]   ###总流入金额
zlc_bs = ls_ic_pay('支出',ls1)[3]   ###总流出金额




ind = ls_in_bn.index
ind_tl = list(ind)


ind1 = ls_out_bn.index
ind_tl1 = list(ind1)
ls_sr_ten = ls_ic_pay('收入',ls1)[1]
ls_sr_ten = ls_sr_ten[['金额','笔数']]        ###前十收入对手方  为剔除关联企业

ls_sr_ten['金额占比'] = ls_sr_ten['金额'].apply(lambda x: x / zlr_je)


ls_zc_ten = ls_ic_pay('支出',ls1)[1]
ls_zc_ten = ls_zc_ten[['金额','笔数']]        ###前十支出对手方  为剔除关联企业
ls_zc_ten['金额占比'] = ls_zc_ten['金额'].apply(lambda x: x / zlc_je)
# loc_qs = len_rjye1 + 4 + 7 + 5


ls_sr_ten.to_excel(result5, sheet_name='银行流水', startrow=loc_qs+2, startcol=2,index=True)
ls_zc_ten.to_excel(result5, sheet_name='银行流水', startrow=loc_qs+16, startcol=2,index=True)



def oneyear():
    nowTime = datetime.datetime.now()

    nowTime = nowTime.strftime('%Y-%m-%d')
    # print(type(nowTime))
    b = pd.date_range(end=nowTime, periods=12, freq='M')
    pydate_array = b.to_pydatetime()
    date_only_array = np.vectorize(lambda s: s.strftime('%Y-%m'))(pydate_array)
    date_only_series = pd.Series(date_only_array)
    ls_bz = pd.DataFrame(columns=date_only_series)
    # ls_bz.fillna(0.0, inplace=True)

    return ls_bz



"""
关联企业流水   以下内容暂时先注释掉 9.13  3：47

"""
print(company_name)
if company_name ==0:
   company_name = '中租（北京）模架工程技术有限公司'

glqy = pd.read_excel(r'/Users/hh/Desktop/2.3银行流水/关联企业-%s.xlsx'%(company_name), skiprows=0)

glqy = glqy['公司名称'].tolist()

ls_name_removed1 = []


def gl(ind_tl, ls_in_bn):  ###流入 ind_tl   流出 ind_tl1
    guanlian_in = []
    # print(ind_tl)
    for i, value in enumerate(ind_tl):

        if value in glqy:
            guanlian_in.append(i)
            ls_name_removed1.append(value)

    ls_in = ls_in_bn.iloc[guanlian_in, :]

    ls_in = ls_in[['金额', '笔数']]  ###如果不需要笔数，从这里删除，对应886行的"笔数"也得删除

    gl_in_sum = ls_in['金额'].sum()  ###关联企业 流入值   第一个总览表有用

    gl_in_count = ls_in['笔数'].sum()

    # ls_chouzi_in = pd.concat([ls_p2p_in, ls_fy_in], axis=0, join='inner')

    return gl_in_sum, gl_in_count, ls_in, ls_name_removed1


gl_in_sum = gl(ind_tl, ls_in_bn)[0]  ###关联流入总额

gl_out_sum = gl(ind_tl1, ls_out_bn)[0]  ###关联流出总额

gl_in_count = gl(ind_tl, ls_in_bn)[1]  ###关联流入总笔数
gl_out_count = gl(ind_tl1, ls_out_bn)[1]  ###关联流出总笔数

ls_name_removed1 = gl(ind_tl, ls_in_bn)[3]

ls_name_removed12 = gl(ind_tl1, ls_out_bn)[3]

ls_gl_in = gl(ind_tl, ls_in_bn)[2]
ls_gl_in['金额占比'] = ls_gl_in['金额'].apply(lambda x: x / zlr_je)


ls_gl_out = gl(ind_tl1, ls_out_bn)[2]

###axis=1需删掉  只对列进行处理
ls_gl_out['金额占比'] = ls_gl_out['金额'].apply(lambda x: x / zlc_je)

glzb_lr = gl_in_sum / zlr_je
fglzb_lr = 1 - glzb_lr
glzb_lc = gl_out_sum / zlc_je
fglzb_lc = 1 - glzb_lc

###表12表12表12.。。。。。。。
gljy = pd.DataFrame({'分类': ['流入', '流入', '流出', '流出'], '项目': ['关联交易', '非关联交易', '关联交易', '非关联交易'], \
                     '金额': [gl_in_sum, zlr_je - gl_in_sum, gl_out_sum, zlc_je - gl_out_sum], \
                     '交易笔数': [gl_in_count, zlr_bs - gl_in_count, gl_out_count, zlc_bs - gl_out_count], \
                     '金额占比': [glzb_lr, fglzb_lr, glzb_lc,fglzb_lc]})

ls_gl_je = pd.DataFrame(columns=['金额', '笔数', '金额占比'])

ls_gl_je = pd.concat([ls_gl_je, ls_gl_in, ls_gl_out], axis=0, join='inner')

ls_gl_je.insert(0, '分类', ['流入'] * len(ls_gl_in.index) + ['流出'] * len(ls_gl_out.index))

ls_gl_je.index.name = '对方户名'

ls_gl_je = ls_gl_je.reset_index(drop=False)  #

ls_gl_je.set_index(['分类', '对方户名'], inplace=True)  ###表13表13表13
# ls_gl_je



ls1_rm = ls1[~ls1['对方户名'].isin(ls_name_removed1)]  ###将关联企业在列表里移除

ls1_rm = ls1_rm[~ls1_rm['对方户名'].isin(ls_name_removed12)]


ls1_rm.reset_index(drop=True, inplace=True)  ###重置索引

len0 = loc_part7+3
len_gap = 2
len_gl = len(ls_gl_je['金额'])

if len_gl>48:
    loc_part7 = 305


gljy.to_excel(result5, sheet_name='银行流水', startrow = loc_part7+2, startcol=1,index=False)
ls_gl_je.to_excel(result5, sheet_name='银行流水', startrow = loc_part7+2, startcol=7,index=True)
ls_gl_je.to_excel(result5, sheet_name='银行流水', startrow = loc_part7+2, startcol=11,index=False)


"""
剔除账号异常退款
"""


def yc(a):  ##a:收入/支出   b:明细  c:附言/摘要

    yc_list1 = []
    yc_list2 = []
    yc_list3 = []
    for i, value in enumerate(ls1_rm['附言']):
        value = str(value)
        if value.find('不符') >= 0 or value.find('异常') >= 0 or value.find('有误') >= 0 or value.find('不存在') >= 0 \
            or value.find('无此账号') >= 0 or value.find('退票重付') >= 0 or value.find('户名误') >= 0 or value.find('账号误') >= 0:
            yc_list1.append(value)
            yc_list2.append(i)

    for i in yc_list2:
        p = ls1_rm.loc[i, '收入']
        b = i - 1

        for j in range(10):
            if ls1_rm.loc[b, '支出'] == p:
                yc_list3.append(b)
                break
            else:
                b -= 1
    yc_list2 = list(set(yc_list3 + yc_list2))

    ls_in = ls1_rm.iloc[yc_list2, :]

    yc_in_sum = ls_in['{0}'.format(a)].sum()

    yc_in_count = ls_in['{0}'.format(a)][ls_in['{0}'.format(a)] > 0].count()

    ls_in['金额'] = ls_in['收入'] + ls_in['支出']
    ls_in = ls_in[['对方户名', '金额']]
    ls_in['笔数'] = 1

    return yc_in_sum, yc_in_count, ls_in, yc_list2


tk_sum_lr = yc('收入')[0]  ###异常流入总额  第一个总览表有用
tk_sum_lc = yc('支出')[0]  ###异常流出总额   第一个总览表有用

tk_count_lr = yc('收入')[1]  ###异常流入总笔数
tk_count_lc = yc('支出')[1]  ###异常流出总笔数

ls_yc = yc('收入')[2]

yc_tk_lr = ls_yc.iloc[:tk_count_lr,:]
yc_tk_lc = ls_yc.iloc[tk_count_lr:,:]

yc_tk_lr['金额占比'] = yc_tk_lr['金额'].apply(lambda x: x / zlr_je)    ###axis=1需删掉  只对列进行处理
yc_tk_lc['金额占比'] = yc_tk_lc['金额'].apply(lambda x: x / zlc_je)

yc_tk_lr.set_index('对方户名', inplace=True)
yc_tk_lc.set_index('对方户名', inplace=True)



yc_list2 = yc('收入')[3]

ls1_rm = ls1_rm[~ls1_rm.index.isin(yc_list2)]


ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引




"""

筹资性流水

"""

"""
类金融机构  

"""

ls_p2p_list = pd.read_excel(r'/Users/hh/Desktop/2.3银行流水/P2P名单.xlsx', skiprows=0)

def wangdai(ls1_rm):
    ind_p2p = []
    # ls_name_list1 = []
    p2p_in = 0
    for i, value in enumerate(ls1_rm['对方户名']):
        for j in ls_p2p_list['平台']:
            value = str(value)          ###报错 'int' object has no attribute 'find'
            if value.find(j) >= 0:
                p2p_in += 1
                ind_p2p.append(i)
                # ls_name_list1.append(value)
                break

    ls_p2p = ls1_rm.iloc[ind_p2p, :]

    p2p_in_sum = ls_p2p['收入'].sum()
    p2p_out_sum = ls_p2p['支出'].sum()
    # p2p_in_count = ls_p2p_in[ls_p2p_in['收入'].values>0].count()
    # p2p_out_count = ls_p2p_in[ls_p2p_in['支出'].values>0].count()
    ls1_rm = ls1_rm[~ls1_rm.index.isin(ind_p2p)]
    ls1_rm = ls1_rm.reset_index(drop=True)
    return p2p_in_sum, p2p_out_sum, ls1_rm, ls_p2p


p2p_in_sum =wangdai(ls1_rm)[0]
p2p_out_sum =wangdai(ls1_rm)[1]



ls_p2p = wangdai(ls1_rm)[3]
ls1_rm = wangdai(ls1_rm)[2]





"""
   非银金融（对方户名匹配）  以下内容暂时先注释掉 9.13  3：47

"""


def feijr(ls1_rm):  ###流入 ind_tl   流出 ind_tl1
    ind_feiyin = []
    # ls_name_list3 = []
    # print(ind_tl)
    for i, value in enumerate(ls1_rm['对方户名']):
        value = str(value)
        if value.find('租赁') >= 0 or value.find('小额贷款') >= 0 or value.find('投资公司') >= 0 \
                or value.find('金融') >= 0 or value.find('保理') >= 0 or value.find('信托') >= 0 or value.find(
            '资产管理') >= 0 or value.find('担保') >= 0 \
                or value.find('典当') >= 0:
            ind_feiyin.append(i)
            # ls_name_list3.append(value)

    # ls_feijr = ls1_rm[ls1_rm['对方户名'].isin(ind_feiyin)]
    ls_feijr = ls1_rm.iloc[ind_feiyin,:]

    feijr_in_sum = ls_feijr['收入'].sum()  ###关联企业 流入值   第一个总览表有用
    feijr_out_sum = ls_feijr['支出'].sum()  ###关联企业 流入值   第一个总览表有用
    ls1_rm = ls1_rm[~ls1_rm.index.isin(ind_feiyin)]
    ls1_rm = ls1_rm.reset_index(drop=True)

    return feijr_in_sum, feijr_out_sum, ls1_rm, ls_feijr


feijr_in_sum = feijr(ls1_rm)[0]
feijr_out_sum = feijr(ls1_rm)[1]

ls_feijr = feijr(ls1_rm)[3]
ls1_rm = feijr(ls1_rm)[2]






# ls_feijr.to_excel(result4, sheet_name='银行流水', startrow=1, startcol=1, index=True)

"""
   银行（对方户名匹配）  以下内容暂时先注释掉 9.13  3：47

"""


def yh(ls1_rm):  ###流入 ind_tl   流出 ind_tl1
    ind_yh = []
    # ls_name_list5 = []
    # print(ind_tl)
    for i, value in enumerate(ls1_rm['对方户名']):
        value = str(value)
        if value.find('银行') >= 0:
            ind_yh.append(i)
            # ls_name_list5.append(value)
    ls_yh = ls1_rm.iloc[ind_yh, :]
    yh_in_sum = ls_yh['收入'].sum()  ###关联企业 流入值   第一个总览表有用
    yh_out_sum = ls_yh['支出'].sum()
    ls1_rm = ls1_rm[~ls1_rm.index.isin(ind_yh)]
    ls1_rm = ls1_rm.reset_index(drop=True)

    return yh_in_sum, yh_out_sum, ls1_rm, ls_yh


yh_in_sum = yh(ls1_rm)[0]
yh_out_sum = yh(ls1_rm)[1]
ls_yh = yh(ls1_rm)[3]
ls1_rm = yh(ls1_rm)[2]


ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引

ls_cz_byname = pd.concat([ls_p2p,ls_feijr, ls_yh], axis=0, join='inner')






"""
  非银机构（借款） 从附言匹配

"""





def rz_byfy(a):  ##a:收入/支出   b:明细  c:附言/摘要

    global ls_rz_lr
    ls_byfy = ls1_rm[ls1_rm['{0}'.format(a)].values > 0]

    # ls_rz_byfy = ls1_rm[ls1_rm['附言'].isin['银总151-手续费']]

    ls_rz_byfy = pd.pivot_table(ls_byfy, values='{0}'.format(a), columns='附言', index='交易时间', aggfunc='sum')

    rz_fy_list = []
    if a=='支出':

        for i, value in enumerate(ls_rz_byfy.columns):  ###前面有累赘写法  需要修改，无需转换成list
        # print(i)
            value = str(value)
            if value.find('手续费YZ') >= 0 or value.find('利息') >= 0 or value.find('还借款') >= 0 or value.find('银行手续费') >= 0 \
                or value.find('手续费') >= 0 or value.find('还款') >= 0 or value.find('还贷款') >= 0 :
                rz_fy_list.append(value)

        ls_rz = ls1_rm[ls1_rm['附言'].isin(rz_fy_list)]
        rz_sum = ls_rz['支出'].sum()

    else:
        for i, value in enumerate(ls_rz_byfy.columns):  ###前面有累赘写法  需要修改，无需转换成list
        # print(i)
            value = str(value)
            if  value.find('借') >= 0:
                rz_fy_list.append(value)

        ls_rz= ls1_rm[ls1_rm['附言'].isin(rz_fy_list)]
        rz_sum = ls_rz['收入'].sum()

    return rz_sum, ls_rz, rz_fy_list

rz_out_sum = rz_byfy('支出')[0]
rz_in_sum = rz_byfy('收入')[0]



ls_rz_lc = rz_byfy('支出')[1]
ls_rz_lr = rz_byfy('收入')[1]


rz_fy_lc = rz_byfy('支出')[2]
rz_fy_lr = rz_byfy('收入')[2]
ls1_rm = ls1_rm[~ls1_rm['附言'].isin(rz_fy_lc)]
ls1_rm = ls1_rm[~ls1_rm['附言'].isin(rz_fy_lr)]
ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引





# ls_rz_lc.to_excel(result4, sheet_name='银行流水', startrow=1, startcol=12, index=True)
"""
筹资性流水汇总
"""



###表7表7表7
# czls = pd.DataFrame({'分类': ['流入', '流入', '流出', '流出'], '项目': ['筹资性流水', '非筹资性流水', '筹资性流水', '非筹资性流水'], \
#                      '金额': [cz_in_sum, fcz_in_sum, cz_out_sum, fcz_out_sum], \
#                      '交易笔数': [cz_in_count, fcz_in_count, cz_out_count, fcz_out_count], \
#                      '金额占比': [czzb_lr, fczzb_lr, czzb_lc,fczzb_lc]})

###表8表8表8
czqd = pd.DataFrame({'分类': ['流入',  '流出'], '银行': [yh_in_sum, yh_out_sum], \
                     '非银金融': [feijr_in_sum+ rz_in_sum, feijr_out_sum + rz_out_sum], \
                     '类金融': [p2p_in_sum , p2p_out_sum]})




czqd.to_excel(result5, sheet_name='银行流水', startrow = loc_part5+13, startcol=1, index=False)

ls_cz_all = pd.concat([ls_rz_lc, ls_rz_lr,ls_cz_byname], axis=0 , join='inner')

lixi_list = []
for i in ls_cz_all['附言']:
    i = str(i)
    if i.find('利息')>=0 or i.find('手续费')>=0:
        lixi_list.append(i)
ls_cz_lixi = ls_cz_all[ls_cz_all['附言'].isin(lixi_list)]

lixi = ls_cz_lixi['支出'].sum()
ls_cz_benjin = ls_cz_all[~ls_cz_all['附言'].isin(lixi_list)]


benjin_lr = ls_cz_benjin['收入'].sum()
benjin_lc = ls_cz_benjin['支出'].sum()

czls = pd.DataFrame({'分类': ['流入', '流出'], '本金': [benjin_lr, benjin_lc], '利息': [0.00, lixi]})
czls.to_excel(result5, sheet_name='银行流水', startrow = loc_part5+3, startcol=1, index=False)

cz_in_sum = benjin_lr
cz_out_sum = benjin_lc + lixi
czzb_lr = cz_in_sum / zlr_je
czzb_lc = cz_out_sum / zlc_je







"""

往来款流水   以下内容暂时先注释掉 9.13  5：46

"""


def wlk(ls1_rm):  ##a:收入/支出   b:明细  c:附言/摘要

    ind_wlk = []
    # ls_name_list5 = []
    # print(ind_tl)
    for i, value in enumerate(ls1_rm['附言']):
        value = str(value)
        if value.find('往来款') >= 0:
            ind_wlk.append(i)
            # ls_name_list5.append(value)
    ls_wlk = ls1_rm.iloc[ind_wlk, :]
    wlk_in_sum = ls_wlk['收入'].sum()  ###关联企业 流入值   第一个总览表有用
    wlk_out_sum = ls_wlk['支出'].sum()
    ls1_rm = ls1_rm[~ls1_rm.index.isin(ind_wlk)]
    ls1_rm = ls1_rm.reset_index(drop=True)

    return wlk_in_sum, wlk_out_sum, ls1_rm, ls_wlk

wlk_in_sum = wlk(ls1_rm)[0]
wlk_out_sum = wlk(ls1_rm)[1]

# print(wlk_in_sum)
# print(wlk_out_sum)
ls_wlk = wlk(ls1_rm)[3]
ls1_rm = wlk(ls1_rm)[2]
# wlk_in_count = ls_wlk[ls_wlk['收入'].values>0].count()
#
# wlk_out_count = ls_wlk[ls_wlk['支出'].values>0].count()

wlkzb_lr = wlk_in_sum / zlr_je
fwlkzb_lr = 1 - wlkzb_lr
wlkzb_lc = wlk_out_sum / zlc_je
fwlkzb_lc = 1 - wlkzb_lc

ls_wlk_lr = ls_wlk[ls_wlk['收入'].values!=0]

wlk_in_count = ls_wlk_lr['收入'].count()
ls_wlk_lr['对方户名'].fillna('未知', inplace=True)

ls_wlk_lc = ls_wlk[ls_wlk['支出'].values != 0]
wlk_out_count = ls_wlk_lc['支出'].count()
ls_wlk_lc['对方户名'].fillna('未知', inplace=True)

if ls_wlk_lr.empty:
    a=0
else:
    ls_wlk_lr = pd.pivot_table(ls_wlk_lr, values='收入', index='对方户名', aggfunc=['sum','count'])
    ls_wlk_lr.columns=['金额','交易笔数']
    ls_wlk_lr['金额占比'] = ls_wlk_lr['金额'].apply(lambda x: x / zlr_je)
    a=len(ls_wlk_lr['金额'])

if ls_wlk_lc.empty:
    b = 0
else:
    ls_wlk_lc = pd.pivot_table(ls_wlk_lc, values='支出', index='对方户名', aggfunc=['sum','count'])
    ls_wlk_lc.columns=['金额','交易笔数']
    ls_wlk_lc['金额占比'] = ls_wlk_lc['金额'].apply(lambda x: x / zlc_je)
    b = len(ls_wlk_lc['金额'])

try:   ###需要斟酌

    ls_wlk = pd.concat([ls_wlk_lr,ls_wlk_lc],axis=0, join='outer')
    ls_wlk= ls_wlk[['金额', '交易笔数', '金额占比']]
    ls_wlk = ls_wlk.reset_index(drop=False)
    ls_wlk.rename(columns={'index': '对方户名'}, inplace=True)
    ls_wlk.insert(0, '分类', ['流入']*a + ['流出']*b)
    ls_wlk.set_index(['分类','对方户名'],inplace=True)     ###表15表15表15
except Exception:

    ls_wlk = pd.DataFrame({'分类': ['流入', '流出'], '对方户名': ['-', '-'], \
                        '金额': ['-', '-'], '交易笔数': ['-', '-'],  '金额占比': ['-', '-']})
    ls_wlk.set_index(['分类','对方户名'],inplace=True)



wlk = pd.DataFrame({'分类': ['流入', '流入', '流出', '流出'], '项目': ['往来款', '非往来款', '往来款', '非往来款'], \
                     '金额': [wlk_in_sum, zlr_je - wlk_in_sum, wlk_out_sum, zlc_je - wlk_out_sum], \
                     '交易笔数': [wlk_in_count, zlr_bs - wlk_in_count, wlk_out_count, zlc_bs - wlk_out_count], \
                     '金额占比': [wlkzb_lr, fwlkzb_lr, wlkzb_lc,fwlkzb_lc]})


loc_wlk0 = loc_part7 + len_gap + max(len_gl+1, 4)+4
loc_wlk1 = loc_wlk0+2
len_wlk = len(ls_wlk['金额'])+1
loc_yc0 = loc_wlk1 + max(len_wlk,5) + len_gap+1

len_qb = len_gl+5+len_wlk
if (len_qb>48) and (len_gl<48):
    loc_wlk0= 305

wlk.to_excel(result5, sheet_name='银行流水', startrow=loc_wlk1, startcol=1, index=False)
ls_wlk.to_excel(result5, sheet_name='银行流水', startrow=loc_wlk1, startcol=7, index=True)
ls_wlk.to_excel(result5, sheet_name='银行流水', startrow=loc_wlk1, startcol=11, index=False)




"""
经营性流出      以下内容暂时先注释掉 9.15  9：46

"""


def jy_lc():  ##a:收入/支出   b:明细  c:附言/摘要

    # start_time = time.time()
    ls_zhichu = ls1_rm[ls1_rm['支出'].values != 0]
    # print(ls1_rm.index)
    # print(ls_zhichu.index)
    # print('bbbu')
    # print(ls_wl_sum)
    xc11 = []
    xc_name = []  ##薪酬
    fy_name = []  ##经营性费用
    zj_name = []  ###租金
    sf_name = []  ##税费
    gz_name = []  ##固定资产
    byj_name = [] ##备用金

    # jy_list = ls_jy_sum.columns.tolist()

    for i, value in enumerate(ls_zhichu['附言']):  ###前面有累赘写法  需要修改，无需转换成list

        value = str(value)
        if value.find('工资') >= 0 or value.find('劳务费') >= 0 or value.find('险') >= 0 or value.find('医疗') >= 0 or \
                value.find('公积金') >= 0 or value.find('差旅') >= 0 or value.find('保安服务费') >= 0 \
                or value.find('生育') >= 0 or value.find('失业') >= 0 or value.find('养老') >= 0 or value.find('保安服务费') >= 0:
            # xc.append(i)
            xc_name.append(i)
            # xc11.append(value)


        elif value.find('电费') >= 0 or value.find('水费') >= 0 or value.find('制冷费') >= 0 or value.find('取暖费') >= 0 \
                or value.find('气费') >= 0 or value.find('水资源费') >= 0 or value.find('天然气') >= 0 or value.find('燃气') >= 0:
            # fy.append(i)
            fy_name.append(i)

        elif value.find('房租') >= 0 or value.find('物业') >= 0 or value.find('仓储费') >= 0 or value.find('房屋定金') >= 0 \
                or value.find('房屋补偿款') >= 0:
            # zj.append(i)
            zj_name.append(i)

        elif value.find('税') >= 0:
            # sf.append(i)
            sf_name.append(i)

        # elif value.find('固资') >= 0 or value.find('设备款') >= 0:
        #     # gz.append(i)
        #     gz_name.append(value)

        elif value.find('备用金') >= 0:
            # gz.append(i)
            byj_name.append(i)


        else:
            pass



    ls_zhichu = ls_zhichu.reset_index(drop=False)  ###重置索引，使他保持连续性reset

    # return xc,fy,zj,sf,gz
    return ls_zhichu, xc_name, fy_name, zj_name, sf_name,  byj_name


gongzi_name = []
ls_byname = ls1_rm[ls1_rm['支出'].values > 0]
# ls_jy_sum = pd.pivot_table(ls_byname, values='支出', columns='附言', index='支出', aggfunc='mean')
for i, value in enumerate(ls1_rm['附言']):
    value = str(value)
    if value.find('工资') >= 0:
        gongzi_name.append(value)
ls_jy_gongzi = ls1_rm[ls1_rm['附言'].isin(gongzi_name)]


ls_jy_gongzi = ls_jy_gongzi[ls_jy_gongzi['支出'].values <20000]
ls_jy_gongzi = ls_jy_gongzi[ls_jy_gongzi['支出'].values >1000]


gongzi = ls_jy_gongzi['支出'].mean()

gongzi = round(gongzi, 2)

if gongzi > 0:

    print('aaaaaaaaaaaaaaaaa')
    print('hhhhhhhhhhh')
    ls_gz = pd.read_excel(r'/Users/hh/Desktop/2.3银行流水/全国各城市工资水平.xlsx', skiprows=0)
    ls_gz.set_index('城市', inplace=True)
    loc_gz = ls_gz.loc['太原市','单月平均工资']

    i = '太原市'
    ls_gz1 = pd.DataFrame({'明细':['公司个人平均工资','%s个人平均工资'%(i)],'工资(元)':[gongzi, loc_gz]})
    ls_gz1.set_index('明细',inplace=True)

    ls_gz1.to_excel(result5, sheet_name='银行流水', startrow = loc_part5-6, startcol=1,index=True)
    ls_gz1.to_excel(result5, sheet_name='银行流水', startrow = loc_part5-6, startcol=3,index=False)

else:
    pass


def jy_out(xc_name, ls_zhichu, a):

    ls_jy_xc = ls_zhichu.iloc[xc_name,:]
    print(ls_jy_xc)


    ls_jy_xc.set_index('index', inplace=True)
    ind_lc = ls_jy_xc.index.to_list()

    # print(ls_jy_xc)
    # print('1234567890')
    ls_jy_xc.set_index('交易时间', inplace=True)
    ls_jy_xc = ls_jy_xc.resample('M', kind='period').sum()

    ls_jy_xc = ls_jy_xc[['支出']]
    ls_jy_xc = ls_jy_xc.T

    # ls_jy_xc1 = pd.DataFrame(ls_jy_xc1)


    ls_jy_xc.insert(0, '营业明细', ['{0}'.format(a)])

    ls_jy_xc.set_index('营业明细', inplace=True)

    ls_jy_xc['金额'] = ls_jy_xc.apply(lambda x: x.sum(), axis=1)


    df_id = ls_jy_xc['金额']                      ###改变列的位置，先复制，后删除，再插入
    ls_jy_xc = ls_jy_xc.drop('金额', axis=1)
    ls_jy_xc.insert(0, '金额', df_id)

    return ls_jy_xc, ind_lc

ls_jy_fy = jy_out(jy_lc()[2], jy_lc()[0], '经营费用')[0]
ls_jy_xc = jy_out(jy_lc()[1], jy_lc()[0], '薪酬')[0]  ###获取薪酬数据
  ###获取水电气数据
print(ls_jy_fy)
print('lllllllllllllllllllllllllllllllllllllll1234')

ls_jy_zj = jy_out(jy_lc()[3], jy_lc()[0], '租金')[0]   ###获取租金数据
ls_jy_sf = jy_out(jy_lc()[4], jy_lc()[0], '税费')[0]   ###获取税收数据
# ls_jy_gz = jy_out(jy_lc()[5], jy_lc()[0], '固定资产')[0]   ###获取固定资产数据
ls_jy_byj = jy_out(jy_lc()[5], jy_lc()[0], '备用金')[0]   ###获取备用金数据


ind_xc= jy_out(jy_lc()[1], jy_lc()[0], '薪酬')[1]
ind_fy= jy_out(jy_lc()[2], jy_lc()[0], '经营费用')[1]
ind_zj= jy_out(jy_lc()[3], jy_lc()[0], '租金')[1]
ind_sf= jy_out(jy_lc()[4], jy_lc()[0], '税费')[1]
ind_byj= jy_out(jy_lc()[5], jy_lc()[0], '备用金')[1]

ls_jy_lc = pd.concat([ls_jy_xc,ls_jy_fy,ls_jy_zj,ls_jy_sf,ls_jy_byj],axis=0,join = 'outer')
ls_jy_lc.fillna(0.0, inplace= True)

# ls_jy_lc.to_excel(result4, sheet_name='银行流水4', index=False)
# result4.save()

ls_name_removed3 = ind_xc + ind_fy + ind_zj + ind_sf + ind_byj

ls1_rm = ls1_rm[~ls1_rm.index.isin(ls_name_removed3)]
# print(ls1_rm11)
ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引，使他保持连续性reset
# print(ls1_rm)


"""
经营性流出（其他费用）以下内容暂时先注释掉 9.15  9：46
"""


def jy_lc_qt():  ##a:收入/支出   b:明细  c:附言/摘要

    ls_zhichu1 = ls1_rm[ls1_rm['支出'].values > 0]

    qt_name = []

    for i, value in enumerate(ls_zhichu1['附言']):  ###前面有累赘写法  需要修改，无需转换成list
        value = str(value)
        if value.find('费') >= 0 or value.find('低耗') >= 0 or value.find('耗材') >= 0 or value.find('押金') >= 0 \
                or value.find('质保金') >= 0 or value.find('退款') >= 0 or value.find('公交款') >= 0 or value.find(
            '退卡') >= 0 or value.find('退保证金') >= 0 \
                or value.find('保洁款') >= 0 or value.find('退货款') >= 0 or value.find('罚款') >= 0 or value.find('航拍款') >= 0 \
                or value.find('违约金') >= 0 or value.find('工程款') >= 0 or value.find('油') >= 0 or value.find('保证金') >= 0 \
                or value.find('招待') >= 0 or value.find('盐') >= 0 or value.find('交通') >= 0 or value.find('个人借款') >= 0 \
                or value.find('维修') >= 0 or value.find('年检') >= 0 or value.find('退现') >= 0 or value.find('汽车租') >= 0 \
                or value.find('赔偿款') >= 0 or value.find('报销') >= 0 or value.find('办公用品') >= 0 or value.find('打印机') >= 0:

            # xc.append(i)
            qt_name.append(i)
        else:
            pass
    ls_zhichu1 = ls_zhichu1.reset_index(drop=False)

    return qt_name, ls_zhichu1

qt_name = jy_lc_qt()[0]
ls_jy_sum1 = jy_lc_qt()[1]

ls_jy_qt = jy_out(qt_name, ls_jy_sum1, '其他')[0]
ind_qt= jy_out(qt_name, ls_jy_sum1, '其他')[1]

ls_jy_lc =pd.concat([ls_jy_lc, ls_jy_qt], axis=0, join='inner')



ls1_rm = ls1_rm[~ls1_rm.index.isin(ind_qt)]  ###移除附言里的融资流水
ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引，使他保持连续性reset

# lr22 = ls1_rm['收入'].sum()
# lc22 = ls1_rm['支出'].sum()
# lr11 = lr11-lr22
# lc11 = lc11-lc22
# print(lr11)
# print(lc11)
# print('a4')



"""
投资性流水

"""


def ls_tz(ls1_rm):

    ls_tz = ls1_rm['附言'].tolist()

    ls_tz_list = []
    ls_tz_i = []

    for i, value in enumerate(ls_tz):
        value = str(value)
        if value.find('理财') >= 0 or value.find('活期') >= 0 or value.find('利息入账') >= 0 \
            or value.find('通知存款') >= 0 or value.find('结息') >= 0 or value.find('购买') >= 0 or value.find('赎回') >= 0:
            ls_tz_list.append(value)

            # print(i)

            ls_tz_i.append(i)
    tz_pdnull = []
    for i in ls_tz_i:
        a = i - 1

        if ls1_rm.loc[i, '对方户名'] == nan:
            for j in range(8):
                if ls1_rm.iloc[a, 0] is nan:
                    a -= 1
                    tz_pdnull.append(a + 1)
                else:
                    pass
    ls_tz_i = list(set(ls_tz_i + tz_pdnull))  ###两个列表合并

    ls_tz1 = ls1_rm.iloc[ls_tz_i, :]

    return ls_tz1, ls_tz_i


ls_tz1 = ls_tz(ls1_rm)[0]  ###获得投资性流水  dataframe
ls_tz_i = ls_tz(ls1_rm)[1]


ls_tz_in = ls_tz1[ls_tz1['收入'] != 0]
ls_tz_out = ls_tz1[ls_tz1['支出'] != 0]

print(ls_tz_in)
print(ls_tz_out)

# ls_tz_in.to_excel(result4, sheet_name='银行流水1', index=True)
# ls_tz_out.to_excel(result4, sheet_name='银行流水2', index=True)
# ls_yhls[2].to_excel(result4, sheet_name='银行流水3', index=True)

tz_in_sum = ls_tz_in['收入'].sum()
tz_in_count = ls_tz_in['收入'].count()

tz_out_sum = ls_tz_out['支出'].sum()
tz_out_count = ls_tz_out['支出'].count()

tzzb_in = tz_in_sum/zlr_je
ftzzb_in = 1 - tzzb_in
tzzb_out = tz_out_sum/zlc_je
ftzzb_out = 1 - tzzb_out

###表9表9表9
tz = pd.DataFrame({'分类': ['流入', '流入', '流出', '流出'], '项目': ['投资性交易', '非投资性交易', '投资性交易', '非投资性交易'], \
                     '金额': [tz_in_sum, zlr_je - tz_in_sum, tz_out_sum, zlc_je - tz_out_sum], \
                     '交易笔数': [tz_in_count, zlr_bs - tz_in_count, tz_out_count, zlc_bs - tz_out_count], \
                     '金额占比': [tzzb_in, ftzzb_in, tzzb_out,ftzzb_out]})

tz1 = tz.iloc[:2,:]
tz2 = tz.iloc[2:,:]

tz1.to_excel(result5, sheet_name='银行流水', startrow = loc_part6+3, startcol=1,index=False)
tz2.to_excel(result5, sheet_name='银行流水', startrow = loc_part6+3, startcol=7,index=False)
# result5.save()

ls1_rm111= ls1_rm[ls1_rm.index.isin(ls_tz_i)]
print(ls1_rm111)

ls1_rm = ls1_rm[~ls1_rm.index.isin(ls_tz_i)]
ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引，使他保持连续性reset

"""
异常交易   排除非工作时间转账 
"""
# myIndex=ls1_rm.index[0<=ls1_rm.index.hour<=8]#
# selectedData=ls1_rm[myIndex]
ls1_rm.set_index('交易时间', inplace=True)

hour = ls1_rm.index.hour
time1 = ls1_rm.index.time

selector = ((22 <= hour) & (hour<= 24)) | ((0 < hour) & (hour <= 7))

ls_yc_time = ls1_rm[selector]
yc_time_list = ls_yc_time.index.to_list()

# ls_yc_time['金额'] = ls_yc_time['收入'] + ls_yc_time['支出']

ls_time_lr = ls_yc_time[ls_yc_time['收入'].values > 0]
ls_time_lr['笔数'] = 1
ls_time_lr['金额占比'] = ls_time_lr['收入'].apply(lambda x: x / zlr_je)

ls_time_lr.rename(columns={'收入': '金额'}, inplace=True)
ls_time_lr = ls_time_lr[['对方户名','金额','笔数','金额占比']]
time_sum_lr = ls_time_lr['金额'].sum()
time_count_lr = ls_time_lr['金额'].count()
ls_time_lr.set_index('对方户名', inplace=True)


ls_time_lc = ls_yc_time[ls_yc_time['支出'].values > 0]
ls_time_lc['笔数'] = 1
ls_time_lc['金额占比'] = ls_time_lc['支出'].apply(lambda x: x / zlc_je)
ls_time_lc.rename(columns={'支出': '金额'}, inplace=True)
ls_time_lc = ls_time_lc[['对方户名','金额','笔数','金额占比']]
ls_time_lc.set_index('对方户名', inplace=True)


time_sum_lc = ls_time_lc['金额'].sum()
time_count_lc = ls_time_lc['金额'].count()


ls1_rm = ls1_rm[~ls1_rm.index.isin(yc_time_list)]
ls1_rm = ls1_rm.reset_index(drop=False)



"""

异常交易   大额高频交易 

"""

ls_hf = ls1_rm

ls_hf['金额'] = ls_hf['收入'] + ls_hf['支出']



ls_hf1 = ls_hf

ls_hf = ls_hf.reset_index(drop=False)
ls_hf['交易时间'] = ls_hf['交易时间'].apply(lambda x: x.strftime('%Y-%m-%d'))



ls_yc_hf = pd.pivot_table(ls_hf, values='金额', index=['交易时间','对方户名'], aggfunc=['sum','count'])       ###数量
ls_yc_hf1 = pd.pivot_table(ls_hf1, values='金额', index=['交易时间','对方户名'], aggfunc=['sum','count'])      ###金额

ls_yc_hf1 = ls_yc_hf1.reset_index(drop=False)
ls_yc_hf1.columns = ['交易时间','对方户名','日总流水','笔数1']
ls_yc_hf1.set_index('交易时间', inplace=True)
ls_yc_hf1 = ls_yc_hf1.resample('D', kind='period').sum()
ls_yc_hf1 = ls_yc_hf1[ls_yc_hf1['笔数1'] >= 5]



yc_ind_list = ls_yc_hf1.index.to_list()
yc_ind_list = list(map(str, yc_ind_list))
ls_yc_hf1.index = yc_ind_list

# ls_yc_hf1.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=17, index=True)

# ls_yc_hf = ls_yc_hf.reset_index(drop=False)
ls_yc_hf.columns = ['金额','笔数']

# ls_yc_hf.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=1, index=True)




ls_yc_hf = ls_yc_hf[ls_yc_hf['笔数'].values >= 5]

# ls_yc_hf.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=6, index=True)



ls_yc_hf = ls_yc_hf.reset_index(drop=False)

# ls_yc_hf.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=11, index=True)

yc_ind_list = ls_yc_hf['交易时间'].to_list()

yc_ind_list = sorted(set(yc_ind_list))

ls_yc_hf1 = ls_yc_hf1.loc[yc_ind_list, :]



ls_hfje = pd.DataFrame()

if ls_yc_hf.empty:
    pass
else:
    ls_hfje = pd.merge(ls_yc_hf, ls_yc_hf1, left_on='交易时间', right_index=True)
    ls_hfje = ls_hfje[ls_hfje['日总流水'] > 5000000]
    ls_hfje['巨额标准'] = ls_hfje['日总流水'].apply(lambda x: float(x*0.2))
    ls_hfje = ls_hfje[ls_hfje['金额'] > ls_hfje['巨额标准']]



"""
下面代码实现将高频巨额从ls1_rm移除
"""
if ls_hfje.empty:
    pass
else:
    yc_hf_list = ls_hfje['对方户名'].to_list()

    yc_hf_list1 = ls_hfje['交易时间'].to_list()

    yc_hf_rm = ls1_rm[ls1_rm['对方户名'].isin(yc_hf_list)]

    yc_hf_rm['交易时间'] = yc_hf_rm['交易时间'].apply(lambda x: x.strftime('%Y-%m-%d'))
    yc_rm_list = yc_hf_rm['交易时间'].to_list()
    yc_rm_list = list(map(str, yc_rm_list))
    yc_hf_rm['交易时间'] = yc_rm_list



    yc_hf_rm = yc_hf_rm[yc_hf_rm['交易时间'].isin(yc_hf_list1)]


    print(yc_hf_rm)

def hfje(a, yc_hf_rm):
    yc_hf_rm = yc_hf_rm[yc_hf_rm['{0}'.format(a)].values > 0]
    yc_hf_rm = pd.pivot_table(yc_hf_rm, values='{0}'.format(a), index='对方户名', aggfunc=['sum', 'count'])
    print('pppppppp')
    print(yc_hf_rm)

    yc_hf_rm.columns = ['金额', '笔数']

    print(yc_hf_rm)
    print('uuuuuuuuuuu')
    hf_sum_lr = yc_hf_rm['金额'].sum()
    hf_count_lr = yc_hf_rm['笔数'].sum()


    return yc_hf_rm, hf_sum_lr,  hf_count_lr


if ls_hfje.empty:
    hf_sum_lr=0
    hf_count_lr=0
    hf_sum_lc = 0
    hf_count_lc = 0
    yc_hf_lr = pd.DataFrame()
    yc_hf_lc = pd.DataFrame()
else:
    try:
        yc_hf_lr = hfje('收入',yc_hf_rm)[0]
        yc_hf_lr['金额占比'] = yc_hf_lr['金额'].apply(lambda x: x/zlr_je)
        hf_sum_lr = hfje('收入',yc_hf_rm)[1]
        hf_count_lr = hfje('收入',yc_hf_rm)[2]
        yc_hf_lr = yc_hf_lr[['金额', '笔数', '金额占比']]        ###明细
    except Exception:
        hf_sum_lr=0
        hf_count_lr=0
        yc_hf_lr = pd.DataFrame()


    try:
        yc_hf_lc = hfje('支出',yc_hf_rm)[0]
        yc_hf_lc['金额占比'] = yc_hf_lc['金额'].apply(lambda x: x/zlc_je)
        hf_sum_lc = hfje('支出',yc_hf_rm)[1]
        hf_count_lc = hfje('支出',yc_hf_rm)[2]
        yc_hf_lc = yc_hf_lc[['金额', '笔数', '金额占比']]         ###明细
    except Exception:
        hf_sum_lc = 0
        hf_count_lc = 0
        yc_hf_lc = pd.DataFrame()


    # ls_yc_hf1.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=8, index=True)
    #
    # ls_yc_hf.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=1, index=True)
    #
    # ls_hfje.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=15, index=True)
    #
    # yc_hf_rm.to_excel(result3, sheet_name='银行流水1112', startrow=0, startcol=25, index=True)




    ls1_rm = ls1_rm[~ls1_rm.index.isin(yc_hf_rm.index)]
    ls1_rm = ls1_rm.reset_index(drop=True)  ###重置索引，使他保持连续性reset





"""
异常交易汇整

"""
yc_sum_lr = tk_sum_lr + time_sum_lr + hf_sum_lr
yc_sum_lc = tk_sum_lc + time_sum_lc + hf_sum_lc

yc_count_lr = tk_count_lr + time_count_lr + hf_count_lr
yc_count_lc = tk_count_lc + time_count_lc + hf_count_lc


yczb_lr = yc_sum_lr / zlr_je
fyczb_lr = 1 - yczb_lr
yczb_lc = yc_sum_lc / zlc_je
fyczb_lc = 1 - yczb_lc

###表16表16表16.。。。。。。。
ycjy = pd.DataFrame({'分类': ['流入', '流入', '流出', '流出'], '项目': ['异常交易', '非异常交易', '异常交易', '非异常交易'], \
                     '金额': [yc_sum_lr, zlr_je - yc_sum_lr, yc_sum_lc, zlc_je - yc_sum_lc], \
                     '交易笔数': [yc_count_lr, zlr_bs - yc_count_lr, yc_count_lc, zlc_bs - yc_count_lc], \
                     '金额占比': [yczb_lr, fyczb_lr, yczb_lc, fyczb_lc]})

ls_yc_lr = pd.concat([yc_tk_lr, ls_time_lr, yc_hf_lr], axis=0, join='outer')
ls_yc_lc = pd.concat([yc_tk_lc, ls_time_lc, yc_hf_lc], axis=0, join='outer')


ls_yc_lr = ls_yc_lr.groupby('对方户名').sum()

ls_yc_lc = ls_yc_lc.groupby('对方户名').sum()
# yc_tk_lr.to_excel(result3 , sheet_name='银行流水111' ,startrow=0, startcol=1,  index=True)
# ls_time_lr.to_excel(result3 , sheet_name='银行流水111' , startrow=0, startcol=8, index=True)
# yc_hf_lr.to_excel(result3 , sheet_name='银行流水111' ,startrow=0, startcol=16, index=True)
# yc_tk_lc.to_excel(result3 , sheet_name='银行流水112' ,startrow=0, startcol=1,  index=True)
# ls_time_lc.to_excel(result3 , sheet_name='银行流水112' , startrow=0, startcol=8, index=True)
# yc_hf_lc.to_excel(result3 , sheet_name='银行流水112' ,startrow=0, startcol=16, index=True)
# result3.save()
# print(absde)

lr_count = ls_yc_lr['金额'].count()
lc_count = ls_yc_lc['金额'].count()

ls_yc = pd.concat([ls_yc_lr, ls_yc_lc], axis=0, join='inner')
# ls_yc.columns = ['金额', '笔数', '金额占比']

ls_yc = ls_yc.reset_index(drop=False)
ls_yc.insert(0, '分类', ['流入'] * lr_count + ['流出'] * lc_count)
ls_yc.set_index(['分类', '对方户名'], inplace=True)  ###表17表17表17

len_yc = len(ls_yc['金额'])

if (len_qb+6+len_yc > 48) and (len_qb < 48):
    loc_yc0 = 305

ycjy.to_excel(result5, sheet_name='银行流水', startrow=loc_yc0+2, startcol=1, index=False)
# ls_yc.to_excel(result5, sheet_name='银行流水', startrow=loc_yc0+2, startcol=7, index=True)
ls_yc.to_excel(result5, sheet_name='银行流水', startrow=loc_yc0+2, startcol=11, index=False)

# ls1_rm.to_excel(result4, sheet_name='银行流水4', index=True, startrow=1,startcol=1)
# result4.save()
# print(ls1_rm)

lr11 = ls1_rm['收入'].sum()
lc11 = ls1_rm['支出'].sum()
"""
经营性流水  销售回款

"""
ls1_rm.set_index('交易时间', inplace=True)
def jingying(a, b,ls1_rm):



    jy_lc = ls1_rm[ls1_rm['{0}'.format(a)].values != 0]

    jy_caigou_sum = ls1_rm['支出'].sum()

    # jy_lr = jy_lr.resample('M', kind='period').sum()
    jy_lc = jy_lc.resample('M', kind='period').sum()

    jy_lc.rename(columns={'金额': '{0}'.format(b)}, inplace=True)
    jy_lc = jy_lc.loc[:,['{0}'.format(b)]].T

    jy_lc.index.name = '经营明细'

    jy_lc['金额'] = jy_lc.apply(lambda x: x.sum(), axis=1)

    df_id1 = jy_lc['金额']  ###改变列的位置，先复制，后删除，再插入
    jy_lc = jy_lc.drop('金额', axis=1)
    jy_lc.insert(0, '金额', df_id1)

    return jy_lc


jy_lc = jingying('支出','采购款',ls1_rm)

ls_jy_lc = pd.concat([jy_lc, ls_jy_lc ], axis=0, join='outer')

ls_jy_lc.loc['总计'] = ls_jy_lc.apply(lambda x: x.sum(), axis=0)

jy_lc_je = ls_jy_lc.loc['总计','金额']

ls_lc_bzt = ls_jy_lc

ls_jy_lc['金额占比'] = ls_jy_lc['金额'].apply(lambda x: x/jy_lc_je)

ls_lc_bzt['金额占比'] = ls_lc_bzt['金额'].apply(lambda x: x/jy_lc_je)

ls_lc_bzt = ls_lc_bzt[['金额占比']]

ls_lc_bzt = ls_lc_bzt.drop('总计')

df_id2 = ls_jy_lc['金额占比']  ###改变列的位置，先复制，后删除，再插入

ls_jy_lc = ls_jy_lc.drop('金额占比', axis=1)
ls_jy_lc.insert(0, '金额占比', df_id2)

ls_jy_lc.index.name = '经营明细'




jy_lr = ls1_rm[ls1_rm['收入'].values != 0]
jy_lr.to_excel(result3, sheet_name='银行流水112', startrow=0, startcol=1, index=True)



xshk = []
for i,value in enumerate(jy_lr['附言']):
    value = str(value)
    if value.find('货款') >= 0 or value.find('营业')>=0 or value.find('售')>=0 or value.find('结算款')>=0 or value.find('转款')>= 0\
        or value.find('ZK') >= 0 or value.find('存现')>= 0 or value.find('结算款')>=0 or value.find('回款')>= 0 or value.find('店')>= 0\
        or value.find('预付款') >= 0 or value.find('工程')>= 0 or value.find('施工')>= 0 or value.find('首付')>= 0 or value.find('设计')>= 0\
        or value.find('材料') >= 0:

        xshk.append(value)

jy_lr1 = jy_lr[~jy_lr['附言'].isin(xshk)]
jy_lr_xs1 = jy_lr[jy_lr['附言'].isin(xshk)]

xshk1 = []
for i,value in enumerate(jy_lr1['对方户名']):
    value = str(value)
    if value.find('备付金')>=0 or value.find('分公司')>=0:
        xshk1.append(value)


jy_lr_qt = jy_lr1[~jy_lr1['对方户名'].isin(xshk1)]
jy_lr_xs2 = jy_lr1[jy_lr1['对方户名'].isin(xshk1)]

jy_lr_xs = pd.concat([jy_lr_xs1,jy_lr_xs2],axis=0,join='inner')


# jy_lr_xs.to_excel(result3, sheet_name='银行流水112', startrow=0, startcol=12, index=True)
# jy_lr_qt.to_excel(result3, sheet_name='银行流水112', startrow=0, startcol=24, index=True)
# result3.save()

jy_lr_xs = jingying('收入','销售回款',jy_lr_xs)

jy_lr_qt = jingying('收入','其他',jy_lr_qt)

# jy_lr = jingying('收入','销售回款')

jy_lr = pd.concat([jy_lr_xs,jy_lr_qt],axis = 0,join='inner')

jy_lr.loc['总计'] = jy_lr.apply(lambda x: x.sum(), axis=0)



jy_lr_je = jy_lr.loc['总计','金额']

ls_lr_bzt = jy_lr

jy_lr['金额占比'] = jy_lr['金额'].apply(lambda x: x/jy_lr_je)


ls_lr_bzt['金额占比'] = ls_lr_bzt['金额'].apply(lambda x: x/jy_lr_je)

ls_lr_bzt = ls_lr_bzt[['金额占比']]

ls_lr_bzt = ls_lr_bzt.drop('总计')

df_id3 = jy_lr['金额占比']  ###改变列的位置，先复制，后删除，再插入
jy_lr = jy_lr.drop('金额占比', axis=1)
jy_lr.insert(0, '金额占比', df_id3)

print(len(jy_lr.columns))
print(type(jy_lr.columns))
print(len(ls_jy_lc.columns))
print('aaaaaaaaaaaa1111111')
ret_list = list(set(jy_lr).union(set(ls_jy_lc)))
# if len(jy_lr.columns) > len(ls_jy_lc.columns):
#     jy_lr.columns=ls_jy_lc.columns
#     jy_lr.fillna(0.0, inplace=True)

ls_jy = pd.concat([jy_lr,ls_jy_lc], axis=0)

# ls_jy.columns=ls_jy_lc.columns

ls_jy.loc['净流入'] = ls_jy.iloc[2, 1:] - ls_jy.iloc[10, 1:]

ls_jy.fillna(0.0,inplace=True)

print(ls_jy_lc)
print(ls_jy)

jy_in_sum = jy_lr.loc['总计','金额']
jy_out_sum = ls_jy_lc.loc['总计','金额']


jyzb_in = jy_in_sum/zlr_je
jyzb_out = jy_out_sum/zlc_je

jy_pic = ls_jy.iloc[[2,10,11],2:].T
jy_pic.columns = ['经营性收入','经营性支出','净流入']
len_jy_pic = len(jy_pic['净流入'])

if len(ls_jy.columns)<=11:
    ls_jy.to_excel(result5, sheet_name='银行流水', startrow=loc_part4 + 45, startcol=2, index=True)
    len_jy = 12
    len_jy1 = len(ls_jy.columns)
    # loc_part5 = loc_part4 + 69
    jy_1=0
else:

    col_len = (len(ls_jy.columns)+1)//2
    ls_jy1 = ls_jy.iloc[:,:col_len]
    ls_jy2 = ls_jy.iloc[:,col_len:]

    ls_jy1.to_excel(result5, sheet_name='银行流水', startrow=loc_part4 + 45, startcol=2, index=True)
    ls_jy2.to_excel(result5, sheet_name='银行流水', startrow=loc_part4 + 58, startcol=2, index=True)
    len_jy = 25
    len_jy1 = len(ls_jy.columns)//2
    jy_1 = len(ls_jy.columns)%2
    # loc_part5 = loc_part4 + 81

ls_lr_bzt.to_excel(result5, sheet_name='银行流水', startrow=loc_part4+27, startcol=3, index=True)
ls_lc_bzt.to_excel(result5, sheet_name='银行流水', startrow=loc_part4+27, startcol=9, index=True)
jy_pic.to_excel(result5, sheet_name='银行流水', startrow=loc_part4+4, startcol=3, index=True)


"""
流水结构总览
"""


lsjg_lr = pd.DataFrame({'分类': ['经营性流水', '筹资性流水', '投资性流水', '关联交易','往来款','异常交易','总计'], \
                     '金额': [jy_in_sum, cz_in_sum, tz_in_sum, gl_in_sum,wlk_in_sum,yc_sum_lr,zlr_je], \
                     '金额占比': [jyzb_in, czzb_lr, tzzb_in, glzb_lr, wlkzb_lr, yczb_lr, 1]})

lsjg_lc = pd.DataFrame({'分类': ['经营性流水', '筹资性流水', '投资性流水', '关联交易','往来款','异常交易','总计'], \
                     '金额': [jy_out_sum, cz_out_sum, tz_out_sum, gl_out_sum,wlk_out_sum,yc_sum_lc,zlc_je], \
                     '金额占比': [jyzb_out, czzb_lc, tzzb_out, glzb_lc, wlkzb_lc, yczb_lc, 1]})

lsjg = pd.concat([lsjg_lr,lsjg_lc],axis = 0, join='inner')


lsjg_zzt = pd.DataFrame({'分类': ['经营性流水', '筹资性流水', '投资性流水', '关联交易','往来款','异常交易'], \
                        '流入': [jyzb_in, czzb_lr, tzzb_in,glzb_lr,wlkzb_lr,yczb_lr],
                         '流出':[jyzb_out, czzb_lc, tzzb_out,glzb_lc,wlkzb_lc,yczb_lc]})

lsjg_bzt = pd.DataFrame({'分类':['流入','流出'],'比例': [zlr_je/(zlr_je + zlc_je),zlc_je/(zlr_je + zlc_je)]})
lsjg_bzt.set_index('分类', inplace=True)
# test_data=[]
# for i in lsjg_bzt.index.values:#获取行号的索引，并对其进行遍历：
#   #根据i来获取每一行指定的数据 并利用to_dict转成字典
#   row_data=lsjg_bzt.ix[i,['比例']].to_dict()
#   test_data.append(row_data)
# print(test_data)

lsjg_bzt.to_excel(result5, sheet_name='银行流水' , startrow = 21, startcol=4,index=True)
# lsjg_bzt.to_excel(result4, sheet_name='银行流水', index=True, startrow=1,startcol=1)

workbook  = result5.book
worksheet = result5.sheets['银行流水']
currency_format = workbook.add_format({'num_format': '0.00%'})
#
chart3 = workbook.add_chart({"type":"pie"})

chart3.add_series({

    "categories":"=银行流水!$E$23:$E$24",
    "values":"=银行流水!$F$23:$F$24",
    'data_labels': {'value': True}
})

chart3.set_size({'width': 330, 'height': 360})


chart3.set_chartarea({
    'border': {'color': 'green'},
    'fill':   {'color': 'white'}
})
chart3.set_style(37)

chart3.set_title({"name":"流入-流出比例"})
worksheet.insert_chart("D14", chart3)
# workbook.close()

# result4.save()


lsjg.to_excel(result5, sheet_name='银行流水' , startrow = 38, startcol=2,index=False)
lsjg_zzt.to_excel(result5, sheet_name='银行流水' , startrow = 41, startcol=8,index=False)

ls_sr_ten.to_excel(result5, sheet_name='银行流水' , startrow=loc_qs+2, startcol=5, index=False)
ls_zc_ten.to_excel(result5, sheet_name='银行流水' , startrow=loc_qs+16, startcol=5,index=False)



result5.save()

"""
excel 绘图

"""


wb = load_workbook(r'9.xlsx')
ws1=wb.active
# sheet = wb.get_active_sheet()



"""
全局格式设置
"""
loc_wme = loc_yc0+2+max(len_yc,4)+6

italic24Font = Font(size=9, italic=False)
for i in range(1,loc_yc0+len_yc+10):
    last_row = ws1[i]
    for each_cell in last_row:
        each_cell.font = italic24Font
        each_cell.alignment = Alignment(horizontal='center', vertical='center')
        each_cell.number_format = '#,##0.00'

for i in range(1,18):
    column = get_column_letter(i)
    col = ws1.column_dimensions[column]
    col.fill = PatternFill("solid", fgColor="FFFFFF")

for i in range(1, loc_wme+4):
    last_row = ws1[i]
    for each_cell in last_row:
        each_cell.fill = PatternFill("solid", fgColor="FFFFFF")


len_lastfoutr = len_gl+len_yc+len_wlk+30

for i in range(len_lastfoutr):
    for item in ws1['E%s:E%s'%(loc_part6,loc_part6+len_lastfoutr)][i]:
        item.number_format = 'general'
    for item in ws1['F%s:F%s'%(loc_part6,loc_part6+len_lastfoutr)][i]:
        item.number_format = '0.00%'


highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=10,color= "ff0100")
highlight.fill = PatternFill("solid", fgColor="DDDDDD")
bd = Side(style='thin', color="000000")
# bd1 = Side(style='thick', color="0000FF")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)


"""
报告标题设置
"""

ws1.merge_cells('A2:N4')
ws1.cell(2, 1).value = '银行流水分析报告'
ws1['A2'].font = Font(name=u'微软雅黑',size=28, bold=True,color=colors.WHITE)
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')

for i in range(1,9):
    last_row = ws1[i]
    for each_cell in last_row:
        each_cell.fill = PatternFill("solid", fgColor="483D8B")

# img0 = Image('logo.png')
#
# newsize = (120, 63)
# img0.width, img0.height = newsize
# ws1.add_image(img0, 'E2')


ws1.merge_cells('A10:D10')
ws1.cell(10, 1).value = "公司名称：%s"%(company_name)
ws1.cell(10, 1).font = Font(name=u'微软雅黑',size=11, bold=True)
ws1.cell(10, 1).alignment = Alignment(horizontal='left', vertical='center')

today=datetime.date.today()
riqi = today.strftime('%Y%m')



bianma_month = datetime.date.today().month
bianma_day = datetime.date.today().day



print(riqi)
print(type(today))

f_out = open(r'num.txt', 'r+')
bianma = f_out.read()
bianma = int(bianma)

print(bianma)

if int(bianma_day) ==1 and bianma not in range(2,100):
    bianma =1
# else:
#
#     bianma = f_out.read()
#     bianma = int(bianma)

ws1.merge_cells('E10:G10')
if bianma < 10:
    ws1.cell(10, 5).value  = "报告编码：ZTMGYHLS%s0%s"%(riqi,bianma)
else:
    ws1.cell(10, 5).value = "报告编码：ZTMGYHLS%s%s" % (riqi, bianma)

ws1.cell(10, 5).font = Font(name=u'微软雅黑',size=11, bold=True)
ws1.cell(10, 5).alignment = Alignment(horizontal='left', vertical='center')



# today=datetime.date.today()
ws1.merge_cells('I10:J10')
ws1.cell(10, 9).value = "报告日期：%s"%(today)
ws1.cell(10, 9).font = Font(name=u'微软雅黑',size=11, bold=True)
ws1.cell(10, 9).alignment = Alignment(horizontal='left', vertical='center')


"""
PART1 银行流水结构总览
"""
for i in range(2):
    for item in ws1['F23:F24'][i]:
        item.number_format = '0.00%'

for i in range(6):
    for item in ws1['J43:K48'][i]:
        item.number_format = '0.00%'          ###柱状图数据转换格式

for i in range(14):
    for item in ws1['E40:E53'][i]:
        item.number_format = '0.00%'
    for item in ws1['D40:D53'][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')


ws1.merge_cells('B40:B46')
ws1.cell(40, 2).value = '流入'
ws1.merge_cells('B47:B53')
ws1.cell(47, 2).value = '流出'

'''
黑框
'''
for i in range(15):

    for item in ws1['B39:E53'][i]:
        item.border = highlight.border
###蓝色填充

for cell in ws1['C39:E39'][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")


"""
PART2 日均余额分析
"""
if rjye_qujian < 12:

    for i in range(2):
        for item in ws1['C62:E63'][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')
        for item in ws1['C89:N90'][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')
else:

    for i in range(2):
        for item in ws1['C62:F63'][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')
        for item in ws1['C89:N90'][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')

if loc_qs >= 94:
    for i in range(2):
        for item in ws1['C92:N93'][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')
else:
    pass

###黑框
if rjye_qujian<12:
    for i in range(3):
        for item in ws1['B61:E63'][i]:
            item.border = highlight.border
else:
    for i in range(3):
        for item in ws1['B61:F63'][i]:
            item.border = highlight.border

len_rj1 = 2+len_rj1
print(len_rj1)

zm1 = get_column_letter(len_rj1)
print(zm1)


if rj_1 ==0:

    for i in range(len_rj):
        for item in ws1['B%s:%s%s'%(len_rjye1+8, zm1, len_rjye1+8 +len_rj)][i]:
            item.border = highlight.border
else:
    zm1 = get_column_letter(len_rj1+1)
    for i in range(len_rj):
        for item in ws1['B%s:%s%s'%(len_rjye1+8, zm1, len_rjye1+8 +len_rj)][i]:
            item.border = highlight.border

###蓝色填充
if rjye_qujian<12:
    for cell in ws1['C%s:E%s'%(len_rjye+1, len_rjye+1)][0]:
        cell.fill = PatternFill("solid", fgColor="B0C4DE")
else:
    for cell in ws1['C%s:F%s' % (len_rjye + 1, len_rjye + 1)][0]:
        cell.fill = PatternFill("solid", fgColor="B0C4DE")

for cell in ws1['C%s:%s%s'%(len_rjye1+8, zm1, len_rjye1+8)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")

if loc_qs>= 96:
    for cell in ws1['C%s:%s%s' % (len_rjye1+11, zm1, len_rjye1+11)][0]:
        cell.fill = PatternFill("solid", fgColor="B0C4DE")
else:
    pass

"""
PART3 收支结构前十对象汇总
"""

for i in range(loc_qs+3, loc_qs+15):
    ws1.merge_cells('C%s:E%s'%(i,i))

for i in range(loc_qs+17, loc_qs+29):
    ws1.merge_cells('C%s:E%s'%(i,i))                 ###合并单元格


for i in range(11):
    for item in ws1['G%s:G%s'%(loc_qs+4, loc_qs+14)][i]:
        item.number_format = 'general'
    for item in ws1['H%s:H%s'%(loc_qs+4, loc_qs+14)][i]:
        item.number_format = '0.00%'
    for item in ws1['G%s:G%s' % (loc_qs+18, loc_qs+28)][i]:
        item.number_format = 'general'
    for item in ws1['H%s:H%s' % (loc_qs+18, loc_qs+28)][i]:
        item.number_format = '0.00%'
    for item in ws1['F%s:F%s' % (loc_qs+4, loc_qs+14)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')
    for item in ws1['F%s:F%s' % (loc_qs+18, loc_qs+28)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')

ws1.merge_cells('B%s:B%s'%(loc_qs+4, loc_qs+14))
ws1.cell(loc_qs+4, 2).value = '流入'
ws1.merge_cells('B%s:B%s'%(loc_qs+18, loc_qs+28))
ws1.cell(loc_qs+18, 2).value = '流出'


for i in range(12):
    for item in ws1['B%s:H%s'%(loc_qs+3, loc_qs+15)][i]:
        item.border = highlight.border

for i in range(12):
    for item in ws1['B%s:H%s'%(loc_qs+17, loc_qs+29)][i]:
        item.border = highlight.border

###蓝色填充
for cell in ws1['C%s:H%s'%(loc_qs+3, loc_qs+3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['C%s:H%s'%(loc_qs+17, loc_qs+17)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")

"""
PART4 经营性流水
"""

for i in range(11):
    for item in ws1['D%s:D%s'%(loc_part4+47 ,loc_part4+57)][i]:
        item.number_format = '0.00%'

for i in range(12):
    for item in ws1['E%s:N%s'%(loc_part4+47 ,loc_part4+58)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')

if loc_part5>=226:
    for i in range(12):
        for item in ws1['D%s:N%s' % (loc_part4+60, loc_part4+71)][i]:
            item.alignment = Alignment(horizontal='right', vertical='center')
else:
    pass


ws1.merge_cells('B%s:B%s'%(loc_part4+46, loc_part4+49))
ws1.cell(loc_part4+46, 2).value = '流入'
ws1.merge_cells('B%s:B%s'%(loc_part4+50, loc_part4+57))
ws1.cell(loc_part4+50, 2).value = '流出'
ws1.merge_cells('B%s:C%s'%(loc_part4+58, loc_part4+58))
ws1.cell(loc_part4+58, 2).value = '净流入'


if loc_part5>226:
    ws1.merge_cells('B%s:B%s' % (loc_part4 + 59, loc_part4 + 62))
    ws1.cell(loc_part4 + 59, 2).value = '流入'
    ws1.merge_cells('B%s:B%s' % (loc_part4 + 63, loc_part4 + 70))
    ws1.cell(loc_part4 + 63, 2).value = '流出'
    ws1.merge_cells('B%s:C%s' % (loc_part4 + 71, loc_part4 + 71))
    ws1.cell(loc_part4 + 71, 2).value = '净流入'
else:
    pass


len_jy1 = 3+len_jy1   ###表示columns从第四列开始写入
zm = get_column_letter(len_jy1)

if jy_1 ==0:

    for i in range(len_jy+1):
        for item in ws1['B%s:%s%s'%(loc_part4+46, zm, loc_part4+47+len_jy)][i]:
            item.border = highlight.border

else:
    zm = get_column_letter(len_jy1+1)
    for i in range(len_jy+1):
        for item in ws1['B%s:%s%s'%(loc_part4+46, zm, loc_part4+47+len_jy)][i]:
            item.border = highlight.border
if gongzi>0:
    for i in range(3):
        for item in ws1['B%s:D%s'%(loc_part5-5,loc_part5-3 )][i]:
            item.border = highlight.border
    ws1.merge_cells('B%s:D%s'%(loc_part5-2, loc_part5-2))
    i='山西省'
    ws1.cell(loc_part5-2, 2).value = '数据来源：%s统计局官网'%(i)
    ws1.cell(loc_part5-2, 2).font = Font(name=u'微软雅黑',size=9,bold=True)
    ws1.cell(loc_part5-2, 2).alignment=Alignment(horizontal='left', vertical='center')
    if gongzi < loc_gz:
        ws1.cell(loc_part5-4,4).font = Font(name=u'微软雅黑', size=10, bold=True, color='008000')
        ws1.cell(loc_part5-3, 4).font = Font(name=u'微软雅黑', size=10, bold=True)
    else:
        ws1.cell(loc_part5 - 4, 4).font = Font(name=u'微软雅黑', size=10, bold=True, color='FF0000')
        ws1.cell(loc_part5 - 3, 4).font = Font(name=u'微软雅黑', size=10, bold=True)
else:
    pass
###蓝色填充


for cell in ws1['C%s:%s%s'%(loc_part4+46, zm, loc_part4+46)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")


if loc_part5>226:
    for cell in ws1['C%s:%s%s' % (loc_part4+59, zm, loc_part4+59)][0]:
        cell.fill = PatternFill("solid", fgColor="B0C4DE")
else:
    pass

for i in range(loc_part5-5, loc_part5-2):
    ws1.merge_cells('B%s:C%s'%(i,i))
# ws1['D%s'%(loc_part5-5)].fill = PatternFill("solid", fgColor="B0C4DE")


"""
PART5 筹资性流水
"""
for i in range(11):
    for item in ws1['E%s:E%s'%(loc_part5,loc_part5+11)][i]:
        item.number_format = 'general'
    for item in ws1['F%s:F%s'%(loc_part5,loc_part5+11)][i]:
        item.number_format = '0.00%'

for i in range(4):
    for item in ws1['D%s:D%s' % (loc_part5 + 5, loc_part5 + 8)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')
    for item in ws1['C%s:E%s' % (loc_part5 + 15, loc_part5 + 18)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')


for i in range(3):
    for item in ws1['B%s:D%s'%(loc_part5+4,loc_part5+7)][i]:
        item.border = highlight.border

for i in range(3):
    for item in ws1['B%s:E%s'%(loc_part5+14,loc_part5+17)][i]:
        item.border = highlight.border


###蓝色填充
for cell in ws1['C%s:D%s'%(loc_part5+4, loc_part5+4)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['C%s:E%s'%(loc_part5+14, loc_part5+14)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")


"""
PART6 投资性流水
"""
for i in range(2):
    for item in ws1['K%s:K%s'%(loc_part6+5,loc_part6+6)][i]:
        item.number_format = 'general'
    for item in ws1['L%s:L%s'%(loc_part6+5,loc_part6+6)][i]:
        item.number_format = '0.00%'
    for item in ws1['D%s:D%s' % (loc_part6 + 5, loc_part6 + 6)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')
    for item in ws1['J%s:J%s' % (loc_part6 + 5, loc_part6 + 6)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')

ws1.merge_cells('B%s:B%s'%(loc_part6+5,loc_part6+6))
ws1.cell(loc_part6+5, 2).value = '流入'
ws1.merge_cells('H%s:H%s'%(loc_part6+5,loc_part6+6))
ws1.cell(loc_part6+5, 8).value = '流出'

for i in range(3):
    for item in ws1['B%s:F%s'%(loc_part6+4,loc_part6+7)][i]:
        item.border = highlight.border
for i in range(3):
    for item in ws1['H%s:L%s'%(loc_part6+4,loc_part6+7)][i]:
        item.border = highlight.border


###蓝色填充
for cell in ws1['C%s:F%s' % (loc_part6 + 4, loc_part6 + 4)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['I%s:L%s' % (loc_part6 + 4, loc_part6 + 4)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")



"""
PART7 关联交易
"""
for i in range(loc_wlk0-loc_part7-3):
    for item in ws1['M%s:M%s'%(loc_part7+4, loc_wlk0)][i]:
        item.number_format = 'general'
    for item in ws1['N%s:N%s'%(loc_part7+4, loc_wlk0)][i]:
        item.number_format = '0.00%'
    for item in ws1['D%s:D%s' %(loc_part7+4, loc_wlk0)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')
    for item in ws1['L%s:L%s' %(loc_part7+4, loc_wlk0)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')

for i in range(loc_part7+3, loc_part7+4+len_gl):       ###I-K列合并单元格
    ws1.merge_cells('I%s:K%s'%(i,i))


ws1.merge_cells('B%s:B%s'%(loc_part7+4, loc_part7+5))
ws1.cell(loc_part7+4, 2).value = '流入'
ws1.merge_cells('B%s:B%s'%(loc_part7+6,loc_part7+7))
ws1.cell(loc_part7+6, 2).value = '流出'

for i in range(5):
    for item in ws1['B%s:F%s'%(loc_part7+3,loc_part7+8)][i]:
        item.border = highlight.border


for i in range(len_gl+1):
    for item in ws1['H%s:N%s'%(loc_part7+3,loc_part7+len_gl+4)][i]:
        item.border = highlight.border

###蓝色填充

for cell in ws1['C%s:F%s' % (loc_part7 + 3, loc_part7 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['I%s:N%s' % (loc_part7 + 3, loc_part7 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")


"""
PART8 往来款
"""

for i in range(loc_yc0-loc_wlk0-3):
    for item in ws1['M%s:M%s'%(loc_wlk0+4, loc_yc0)][i]:
        item.number_format = 'general'
    for item in ws1['N%s:N%s'%(loc_wlk0+4, loc_yc0)][i]:
        item.number_format = '0.00%'
    for item in ws1['D%s:D%s' %(loc_wlk0+4, loc_yc0)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')
    for item in ws1['L%s:L%s' %(loc_wlk0+4, loc_yc0)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')


for i in range(loc_wlk0+3, loc_wlk0+3+len_wlk):       ###I-K列合并单元格
    ws1.merge_cells('I%s:K%s'%(i,i))


ws1.merge_cells('B%s:B%s'%(loc_wlk0+4, loc_wlk0+5))
ws1.cell(loc_wlk0+4, 2).value = '流入'
ws1.merge_cells('B%s:B%s'%(loc_wlk0+6,loc_wlk0+7))
ws1.cell(loc_wlk0+6, 2).value = '流出'

for i in range(5):
    for item in ws1['B%s:F%s'%(loc_wlk0+3,loc_wlk0+8)][i]:
        item.border = highlight.border

for i in range(len_wlk):
    for item in ws1['H%s:N%s'%(loc_wlk0+3, loc_wlk0+len_wlk+3)][i]:
        item.border = highlight.border

###蓝色填充

for cell in ws1['C%s:F%s' % (loc_wlk0 + 3, loc_wlk0 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['I%s:N%s' % (loc_wlk0 + 3, loc_wlk0 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")


"""
PART9 异常交易
"""
for i in range(len_yc+1):
    for item in ws1['M%s:M%s'% (loc_yc0+4,loc_yc0+len_yc+4)][i]:
        item.number_format = 'general'
    for item in ws1['N%s:N%s'% (loc_yc0+4,loc_yc0+len_yc+4)][i]:
        item.number_format = '0.00%'
    for item in ws1['L%s:L%s' % (loc_yc0+4,loc_yc0+len_yc+4)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')


for i in range(loc_yc0+3, loc_yc0+4+len_yc):             ###I-K列合并单元格
    ws1.merge_cells('I%s:K%s'%(i,i))

for i in range(4):
    for item in ws1['D%s:D%s' % (loc_yc0+4,loc_yc0+7)][i]:
        item.alignment = Alignment(horizontal='right', vertical='center')


ws1.merge_cells('B%s:B%s'%(loc_yc0+4, loc_yc0+5))
ws1.cell(loc_yc0+4, 2).value = '流入'
ws1.merge_cells('B%s:B%s'%(loc_yc0+6,loc_yc0+7))
ws1.cell(loc_yc0+6, 2).value = '流出'


for i in range(5):
    for item in ws1['B%s:F%s'%(loc_yc0+3, loc_yc0+8)][i]:
        item.border = highlight.border

for i in range(len_yc+1):
    for item in ws1['H%s:N%s'%(loc_yc0+3, loc_yc0+len_yc+4)][i]:
        item.border = highlight.border


###蓝色填充

for cell in ws1['C%s:F%s' % (loc_yc0 + 3, loc_yc0 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")
for cell in ws1['I%s:N%s' % (loc_yc0 + 3, loc_yc0 + 3)][0]:
    cell.fill = PatternFill("solid", fgColor="B0C4DE")

"""
PART10 联系我们
"""

for i in range(loc_wme+10):
    ws1.row_dimensions[i].height = 12


ws1.merge_cells('A%s:D%s'%(loc_wme, loc_wme))
ws1.cell(loc_wme, 1).value = '联系我们：'
ws1.cell(loc_wme, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wme, 1).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A%s:D%s'%(loc_wme+1, loc_wme+1))
ws1.cell(loc_wme+1, 1).value = '公司名称：中投摩根信息技术（北京）有限责任公司'
ws1.cell(loc_wme+1, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wme+1, 1).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A%s:D%s'%(loc_wme+2, loc_wme+2))
ws1.cell(loc_wme+2, 1).value = '公司地址：北京市海淀区西金大厦601室'
ws1.cell(loc_wme+2, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wme+2, 1).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A%s:D%s'%(loc_wme+3, loc_wme+3))
ws1.cell(loc_wme+3, 1).value = '联系人：    赵林煜'
ws1.cell(loc_wme+3, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wme+3, 1).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A%s:D%s'%(loc_wme+4, loc_wme+4))
ws1.cell(loc_wme+4, 1).value = '联系方式：13641098490'
ws1.cell(loc_wme+4, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wme+4, 1).alignment = Alignment(horizontal='left', vertical='center')


"""
表头格式设置
"""

ws1.merge_cells('D35:H35')
ws1.cell(35, 4).value = '图1.流入/流出占比'

ws1.merge_cells('B38:E38')
ws1.cell(38, 2).value = '表1.银行流水结构总览表'

ws1.merge_cells('H56:L56')
ws1.cell(56, 8).value = '图2.银行流水结构比例'

ws1.merge_cells('B60:F60')
ws1.cell(60, 2).value = '表2.期间日均余额'

ws1.merge_cells('B85:E85')
ws1.cell(85, 2).value = '图3.期间日均余额柱状图'

ws1.merge_cells('H85:L85')
ws1.cell(85, 8).value = '图4.月度日均余额柱状图'

zm2 = get_column_letter(len_rj1)

ws1.merge_cells('B87:%s87'%(zm2))
ws1.cell(87, 2).value = '表3.月度日均余额'

ws1.merge_cells('B%s:H%s'%(loc_qs+2,loc_qs+2))
ws1.cell(loc_qs+2, 2).value = '表4.收入前十对象'

ws1.merge_cells('B%s:H%s'%(loc_qs+16,loc_qs+16))
ws1.cell(loc_qs+16, 2).value = '表5.支出前十对象'


ws1.merge_cells('B%s:K%s'%(loc_part4+22,loc_part4+22))
ws1.cell(loc_part4+22, 2).value = '图5.经营性流水金额-月度'

ws1.merge_cells('B%s:E%s'%(loc_part4+43,loc_part4+43))
ws1.cell(loc_part4+43, 2).value = '图6.营业性收入明细'

ws1.merge_cells('H%s:K%s'%(loc_part4+43,loc_part4+43))
ws1.cell(loc_part4+43, 8).value = '图7.营业性支出明细'

ws1.merge_cells('B%s:%s%s'%(loc_part4+45,zm,loc_part4+45))
ws1.cell(loc_part4+45, 2).value = '表6.经营性流水明细-月度数据'

# ws1.merge_cells('B%s:D%s'%(loc_part5-6,loc_part5-6))
# ws1.cell(loc_part5-6, 2).value = '表7.公司工资与当地工资比较'

ws1.merge_cells('B%s:D%s'%(loc_part5+3,loc_part5+3))
ws1.cell(loc_part5+3, 2).value = '表8.筹资性交易流入-流出'

ws1.merge_cells('B%s:E%s'%(loc_part5+13,loc_part5+13))
ws1.cell(loc_part5+13, 2).value = '表9.筹资性交易渠道'

ws1.merge_cells('H%s:L%s'%(loc_part5+20,loc_part5+20))
ws1.cell(loc_part5+20, 8).value = '图8.筹资性交易渠道'

ws1.merge_cells('B%s:F%s'%(loc_part6+3,loc_part6+3))
ws1.cell(loc_part6+3, 2).value = '表10.投资性交易流入数据'

ws1.merge_cells('H%s:L%s'%(loc_part6+3,loc_part6+3))
ws1.cell(loc_part6+3, 8).value = '表11.投资性交易流出数据'

ws1.merge_cells('B%s:F%s'%(loc_part7+2,loc_part7+2))
ws1.cell(loc_part7+2, 2).value = '表12.关联交易流入-流出'

ws1.merge_cells('H%s:N%s'%(loc_part7+2,loc_part7+2))
ws1.cell(loc_part7+2, 8).value = '表13.关联交易明细'

ws1.merge_cells('B%s:F%s'%(loc_wlk0+2,loc_wlk0+2))
ws1.cell(loc_wlk0+2, 2).value = '表14.往来款流入-流出'

ws1.merge_cells('H%s:N%s'%(loc_wlk0+2,loc_wlk0+2))
ws1.cell(loc_wlk0+2, 8).value = '表15.往来款交易明细'

ws1.merge_cells('B%s:F%s'%(loc_yc0+2,loc_yc0+2))
ws1.cell(loc_yc0+2, 2).value = '表16.异常交易流入-流出'

ws1.merge_cells('H%s:N%s'%(loc_yc0+2,loc_yc0+2))
ws1.cell(loc_yc0+2, 8).value = '表17.异常交易明细'


"""
PART 格式设置
"""
for i in [12,58,loc_qs,loc_part4,loc_part5,loc_part6,loc_part7,loc_wlk0, loc_yc0]:
    last_row = ws1[i]
    for each_cell in last_row:
        each_cell.fill = PatternFill("solid", fgColor="F0E68C")


ws1.merge_cells('A12:B12')
ws1.cell(12, 1).value = 'PART1：银行流水结构总览'
ws1.cell(12, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(12, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A58:B58')
ws1.cell(58, 1).value = 'PART2：日均余额分析'
ws1.cell(58, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(58, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:C%s'%(loc_qs,loc_qs))
ws1.cell(loc_qs, 1).value = 'PART3：收支结构前十对象汇总'
ws1.cell(loc_qs, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_qs, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:B%s'%(loc_part4, loc_part4))
ws1.cell(loc_part4, 1).value = 'PART4：经营性流水'
ws1.cell(loc_part4, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_part4, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:B%s'%(loc_part5, loc_part5))
ws1.cell(loc_part5, 1).value = 'PART5：筹资性流水'
ws1.cell(loc_part5, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_part5, 1).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('A%s:B%s'%(loc_part6, loc_part6))
ws1.cell(loc_part6, 1).value = 'PART6：投资性流水'
ws1.cell(loc_part6, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_part6, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:B%s'%(loc_part7, loc_part7))
ws1.cell(loc_part7, 1).value = 'PART7：关联交易'
ws1.cell(loc_part7, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_part7, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:B%s'%(loc_wlk0, loc_wlk0))
ws1.cell(loc_wlk0, 1).value = 'PART8：往来款'
ws1.cell(loc_wlk0, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_wlk0, 1).alignment = Alignment(horizontal='left', vertical='center')


ws1.merge_cells('A%s:B%s'%(loc_yc0, loc_yc0))
ws1.cell(loc_yc0, 1).value = 'PART9：异常交易'
ws1.cell(loc_yc0, 1).font = Font(name=u'微软雅黑',size=9, bold=True)
ws1.cell(loc_yc0, 1).alignment = Alignment(horizontal='left', vertical='center')



# .graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
"""
图1 流入-流出饼图
"""


# pie = PieChart3D()
# pie.title = "流入-流出比例"
# labels = Reference(ws1, min_col=5, min_row=23, max_row=24)
# data2 = Reference(ws1, min_col=6, min_row=22, max_row=24)
# pie.add_data(data2, titles_from_data=True)
# pie.set_categories(labels)
# pie.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
# ws1.add_chart(pie, "D15")

"""
图2   银行流水结构比例
"""
chart0 = BarChart()      ###LineChart()
chart0.type = "col"
chart0.style = 18
chart0.title = "银行流水结构比例"

data0 = Reference(ws1, min_col=10, max_col=11, min_row=42, max_row=48)
cats0 = Reference(ws1, min_col=9, max_col=9, min_row=43, max_row=48)
chart0.add_data(data0, titles_from_data=True)

chart0.set_categories(cats0)
chart0.shape = 50
# chart0.grouping = "standard"
chart0.width = 15
chart0.height = 8

chart0.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(chart0, "H35")

"""
图3  期间日均余额
"""
c1 = BarChart()      ###LineChart()
c1.type = "col"
c1.style = 18

c1.y_axis.title = '日均余额'
# c1.x_axis.title = '日均余额'
if rjye_qujian<12:
    data = Reference(ws1, min_col=2, max_col=5, min_row=62, max_row=62)
    cats2 = Reference(ws1, min_col=3, max_col=5, min_row=61, max_row=61)
else:
    data = Reference(ws1, min_col=2, max_col=6, min_row=62, max_row=62)
    cats2 = Reference(ws1, min_col=3, max_col=6, min_row=61, max_row=61)

c1.add_data(data, titles_from_data=True, from_rows=True)
# chart1.add_data(data11, titles_from_data=True)

c1.set_categories(cats2)
c1.y_axis.axId = 200
c1.legend.position = 'b'

c2 = LineChart()

c1.title = "期间日均余额（单位：元）"
c2.style = 18
# c2.type = "col"
if rjye_qujian<12:
    data11 = Reference(ws1, min_col=2, max_col=5, min_row=63, max_row=63)
else:
    data11 = Reference(ws1, min_col=2, max_col=6, min_row=63, max_row=63)
c2.add_data(data11, titles_from_data=True, from_rows=True)
c2.set_categories(cats2)
c2.y_axis.majorGridlines = None
c2.y_axis.title = '最小余额'
c2.y_axis.crosses = "max"
s4 = c2.series[0]
s4.marker.symbol = "triangle"

c1 += c2
c1.shape = 8
c1.width = 13
c1.height = 7

c1.smooth = True
c2.smooth = True

c1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(c1, "B65")

"""
图4 每月日均余额
"""
c3 = BarChart()      ###LineChart()
c3.type = "col"
c3.style = 18
c3.title = "月度日均余额（单位：元）"
c3.y_axis.title = '日均余额'
# c3.x_axis.title = '日均余额'

d1 = Reference(ws1, min_col=10, max_col=10, min_row=64, max_row=64+len_ye)
cats1 = Reference(ws1, min_col=9, max_col=9, min_row=65, max_row=64+len_ye)
c3.add_data(d1, titles_from_data=True)
c3.set_categories(cats1)
c3.y_axis.axId = 200
c3.legend.position = 'b'


c4 = LineChart()
c4.title = "月度日均余额"
c4.style = 18
# c2.type = "col"
d2 = Reference(ws1, min_col=11, max_col=11, min_row=64, max_row=64+len_ye)
c4.add_data(d2, titles_from_data=True)
c4.set_categories(cats1)
c4.y_axis.majorGridlines = None
c4.y_axis.title = '最小余额'
c4.y_axis.crosses = "max"
s5 = c4.series[0]
s5.marker.symbol = "triangle"

c3 += c4
c3.shape = 4
c3.width = 15
c3.height = 8
c3.smooth = True
c4.smooth = True
c3.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(c3, "H63")


"""
图5    经营流水折线图
"""
c0 = LineChart()      ###LineChart()
c0.type = "col"
c0.style = 18
c0.title = "经营流水折线图（单位：元）"
c0.y_axis.title = '营业性收入/支出'

d0 = Reference(ws1, min_col=5, max_col=6, min_row=loc_part4+5, max_row=loc_part4+5+len_jy_pic)
cat0 = Reference(ws1, min_col=4, max_col=4, min_row=loc_part4+6, max_row=loc_part4+5+len_jy_pic)
c0.add_data(d0, titles_from_data=True)
c0.y_axis.axId = 200
c0.set_categories(cat0)
c0.shape = 2
# chart0.grouping = "standard"

s1 = c0.series[0]
s2 = c0.series[1]
s2.smooth = True
s1.smooth = True

s1.marker.symbol = "diamond"
s2.marker.symbol = "square"

c01 = LineChart()      ###LineChart()
c01.type = "col"
c0.style = 18
c01.title = "经营流水折线图"
c01.y_axis.title = '净流入'

d0 = Reference(ws1, min_col=7, max_col=7, min_row=loc_part4+5, max_row=loc_part4+5+len_jy_pic)
c01.add_data(d0, titles_from_data=True)
c01.set_categories(cat0)
c01.y_axis.majorGridlines = None
c01.y_axis.crosses = "max"
c0.x_axis.crosses = "min"

s3 = c01.series[0]
s3.marker.symbol = "triangle"
s3.smooth = True

c0 += c01

c0.width = 30
c0.height = 8

c0.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(c0, "B"+str(loc_part4+1))

"""
图6 经营性流入饼状图

"""

pie2 = PieChart()
pie2.title = "收入饼状图"
labels1 = Reference(ws1, min_col=4, min_row=loc_part4+29, max_row=loc_part4+30)
data22 = Reference(ws1, min_col=5, min_row=loc_part4+29, max_row=loc_part4+30)
pie2.add_data(data22, titles_from_data=False)
pie2.set_categories(labels1)

# pie2.show_hidden_data()

pie2.width = 13
pie2.height = 7

pie2.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(pie2, "B"+str(loc_part4+24))
"""
图7 经营性流出饼状图
"""

pie1 = ProjectedPieChart()
pie1.title = "支出饼状图"
labels = Reference(ws1, min_col=10, min_row=loc_part4+29, max_row=loc_part4+35)
data21 = Reference(ws1, min_col=11, min_row=loc_part4+29, max_row=loc_part4+35)
pie1.add_data(data21, titles_from_data=False)
pie1.set_categories(labels)

pie1.width = 14
pie1.height = 7

pie1.type = "bar"
pie1.splitType = 'percent' # split by position

pie1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))    #去图标边框
ws1.add_chart(pie1, "H"+str(loc_part4+24))


"""
图8   筹资渠道
"""
chart4 = BarChart()      ###LineChart()
chart4.type = "col"
chart4.style = 18
chart4.title = "筹资渠道（单位：元）"


data4 = Reference(ws1, min_col=2, max_col=5, min_row=loc_part5+15, max_row=loc_part5+16)
cats4 = Reference(ws1, min_col=3, max_col=5, min_row=loc_part5+14, max_row=loc_part5+14)
chart4.add_data(data4, titles_from_data=True, from_rows=True)
chart4.set_categories(cats4)
chart4.shape = 40

chart4.width = 13
chart4.height = 7

chart4.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
ws1.add_chart(chart4, "H"+str(loc_part5+1))


"""
自动调整表格宽度
"""
for col in ws1.columns:
    max_length = 0
    column = col[0].column # Get the column name
    column = get_column_letter(column)   ###补充这一句
    for cell in col:
        if cell.coordinate in ws1.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 3) * 1.4
    if adjusted_width >11.7:
        ws1.column_dimensions[column].width = adjusted_width
    else:
        ws1.column_dimensions[column].width = 13.5


bianma = bianma+1
f_out.seek(0)#清除内容
f_out.truncate()
f_out.write(str(bianma))
f_out.close()

wb.save(r'%s-银行流水报告.xlsx'%(company_name))

end_time = time.time()
run_time = end_time - start_time
print('run_time:',run_time)
